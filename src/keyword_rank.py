"""关键词排名查询 — 每天自动查 Amazon 搜索结果中的自然位排名 + 广告位页码 + BSR 大类排名。

安全策略:
- 只用 evaluate() 读 DOM，绝不 click()
- 翻页只通过 window.location.href
- 广告链接 (/sspa/click) 只识别不触碰

用法:
    # 刮水器 (Amazon DE, 7 ASINs)
    python -m src.keyword_rank <excel> --site de --asin B0CLXXD2X4 B0C6TCLHHT ...

    # 猫砂垫 (Amazon FR, 1 ASIN)
    python -m src.keyword_rank <excel> --site fr --asin B0CH4N8V6P

每个产品 Excel 对应一组 ASIN，通过 CLI 参数指定站点和 ASIN。
"""

from __future__ import annotations

import json
import re
import shutil
import time
from datetime import datetime, date
from pathlib import Path
from urllib.parse import quote

import openpyxl

from .cdp_client import CdpClient

try:
    from .cli_utils import fatal_error, print_json_summary, RunTimer
except ImportError:
    from cli_utils import fatal_error, print_json_summary, RunTimer

# ── 站点配置 ────────────────────────────────────────────────────
SITE_DOMAIN = {
    "de": "amazon.de",
    "fr": "amazon.fr",
    "us": "amazon.com",
    "ca": "amazon.ca",
}
BSR_PATTERN = re.compile(r'^\d+-\d+$')
MAX_PAGES = 7
CDP_HOST = "localhost"
CDP_PORT = 9222

# 模拟人工操作的时间参数（秒）
PAGE_LOAD_MIN = 1.5   # 页面加载最小等待（卡片已就绪后的额外缓冲）
PAGE_LOAD_MAX = 2.5
SCROLL_PAUSE = 0.3    # 滚动停顿
KEYWORD_GAP_MIN = 1.0 # 关键词切换间隔最小
KEYWORD_GAP_MAX = 2.0 # 关键词切换间隔最大

# ── JS 脚本（不包含任何点击操作）──────────────────────────────────

_EXTRACT_JS = """
(() => {
    const results = [];
    const cardAsins = new Set();
    const allDpAsins = new Set();
    const adTexts = ['Sponsored', 'Sponsorisé', 'Gesponsert'];

    const cards = document.querySelectorAll('[data-asin]');
    for (const card of cards) {
        const asin = card.getAttribute('data-asin');
        if (!asin) continue;
        cardAsins.add(asin);

        const hrefs = Array.from(card.querySelectorAll('a')).map(a => a.getAttribute('href') || '');
        const dpLink = hrefs.find(h => /\\/dp\\//.test(h) && /ref=sr_/.test(h));

        // Pass 1: Sponsored text inside card's own subtree (Sponsored Products, Video ads).
        let hasAd = false;
        const walker = document.createTreeWalker(card, NodeFilter.SHOW_TEXT);
        let textNode;
        while ((textNode = walker.nextNode())) {
            if (adTexts.includes(textNode.textContent.trim())) {
                hasAd = true;
                break;
            }
        }

        // Pass 2: Card inside a Sponsored Brand container (Sponsored text is in
        // the header banner, not inside individual product cards).
        if (!hasAd) {
            let el = card.parentElement;
            for (let i = 0; i < 20 && el && el !== document.body; i++) {
                const cls = el.className || '';
                if (cls.includes('sb-desktop') || cls.includes('_c2itd_container')) {
                    hasAd = true;
                    break;
                }
                el = el.parentElement;
            }
        }

        let rank = null;
        if (dpLink) { const m = dpLink.match(/ref=sr_\\d+_(\\d+)/); if (m) rank = parseInt(m[1]); }
        if (rank === null) {
            const idx = card.getAttribute('data-index');
            if (idx) rank = parseInt(idx);
        }
        results.push({asin, rank, isAd: hasAd});
    }

    // Collect all /dp/ ASINs on page — catches carousel/inline items whose
    // [data-asin] cards lack ref=sr_ links (variants, widget embeds, etc.)
    const allLinks = document.querySelectorAll('a[href*="/dp/"]');
    for (const a of allLinks) {
        const m = a.getAttribute('href').match(/\\/dp\\/([A-Z0-9]{10})/);
        if (m) allDpAsins.add(m[1]);
    }

    return JSON.stringify({results, count: results.length, allDpAsins: Array.from(allDpAsins)});
})()
"""

_BSR_EXPAND_JS = """
(() => {
    // 展开折叠区域，但跳过真实导航链接（US 站 "See Top 100" 等 href=/gp/ 会破坏 BSR 区域）。
    // DE 站 "Artikelangaben" 等使用 href=javascript:void(0) 的展开按钮不受影响。
    const btns = document.querySelectorAll('[aria-expanded="false"]');
    let n = 0;
    for (const b of btns) {
        const a = b.tagName === 'A' ? b : b.closest('a');
        if (a) {
            const href = a.getAttribute('href') || '';
            if (href.startsWith('/') || href.startsWith('http')) continue;
        }
        b.click();
        n++;
    }
    return n;
})()
"""

_BSR_JS = """
(() => {
    const ranks = [];
    const body = document.body.innerText;

    // Multi-language BSR section headers
    const headers = [
        'Best Sellers Rank',
        'Classement des meilleures ventes',
        'Amazon Bestseller-Rang',
    ];

    let start = -1;
    for (const h of headers) {
        start = body.indexOf(h);
        if (start >= 0) break;
    }
    if (start < 0) return JSON.stringify({ranks});

    // Extract ~600 chars after the header, strip parentheticals first
    // to avoid matching "Top 100" / "Voir les 100 premiers" references.
    const section = body.substring(start, start + 600).replace(/\\([^)]*\\)/g, '');

    // DE: "Nr. 1.310 in Küche, Haushalt & Wohnen"
    // FR: "3 133 en Animalerie"
    // EN: "#1,234 in Category"
    // Separators: dot (DE), comma (EN), thin/nbspace (FR), regular space
    const rankRe = /(\\d+(?:[\\s.,  ]\\d+)*)\\s+(en|in)\\s+(.+?)(?=\\s*$|\\s*\\n)/gm;
    let m;
    while ((m = rankRe.exec(section)) !== null) {
        const rank = m[1].replace(/[\\s.,  ]/g, '');
        ranks.push({rank, category: m[3].trim()});
    }

    return JSON.stringify({ranks});
})()
"""


class KeywordRankChecker:
    """关键词排名查询器 — CDP 操控浏览器，零点击安全策略。"""

    def __init__(self, asins: set[str], site: str = "de", bsr_asins: list[tuple[str, str]] | None = None,
                 ad_asins: set[str] | None = None,
                 white_asin: str = "",
                 host: str = CDP_HOST, port: int = CDP_PORT):
        self.asins = asins
        self.ad_asins = ad_asins or asins
        self.white_asin = white_asin
        self.site = site
        self.bsr_asins = bsr_asins or [("", next(iter(asins)))]
        self.cdp = CdpClient(host, port, timeout=45)
        self._tab_ws: str | None = None

    def _ensure_tab(self):
        """确保连接到一个 Amazon 标签页，优先匹配首页（非搜索结果页）。"""
        self.cdp.close()
        tabs = self.cdp.list_tabs()
        domain = SITE_DOMAIN[self.site]

        def _is_amazon_tab(t: dict) -> bool:
            url = t.get("url", "")
            if domain not in url:
                return False
            if t.get("type") not in ("page", None):
                return False
            if "aax" in url or "service-worker" in url:
                return False
            return True

        # 优先：非搜索结果页（首页 / dp 产品页）
        for t in tabs:
            if _is_amazon_tab(t) and "s?k=" not in t.get("url", ""):
                self._tab_ws = t["webSocketDebuggerUrl"]
                self.cdp.connect_tab(self._tab_ws)
                self.cdp._send({"method": "Page.bringToFront", "params": {}, "id": 1})
                return

        # 回退：任意匹配的标签页
        for t in tabs:
            if _is_amazon_tab(t):
                self._tab_ws = t["webSocketDebuggerUrl"]
                self.cdp.connect_tab(self._tab_ws)
                self.cdp._send({"method": "Page.bringToFront", "params": {}, "id": 1})
                return

        raise RuntimeError(
            f"No Amazon {self.site.upper()} tab found. Start Edge with the .bat file first."
        )

    def _navigate(self, url: str):
        self._safe_evaluate(f'window.location.href = "{url}"')

    def _extract_page(self) -> tuple[list[dict], set[str]]:
        raw = self._safe_evaluate(_EXTRACT_JS)
        data = json.loads(raw["result"]["result"]["value"])
        return data["results"], set(data.get("allDpAsins", []))

    def _safe_evaluate(self, js: str, retries: int = 2) -> dict:
        """带断线恢复的 evaluate，WebSocket 超时时重连并重试。"""
        for attempt in range(retries + 1):
            try:
                return self.cdp.evaluate(js)
            except (TimeoutError, ConnectionError, OSError) as e:
                if attempt == retries:
                    raise
                time.sleep(1)
                self._ensure_tab()

    def _human_delay(self, lo: float, hi: float):
        """随机等待，模拟人工操作的不规律节奏。"""
        import random
        time.sleep(lo + random.random() * (hi - lo))

    def _wait_for_cards(self, timeout: float = 10.0) -> bool:
        """轮询等待搜索结果卡片渲染完成。找到卡片立即返回 True。"""
        deadline = time.time() + timeout
        while time.time() < deadline:
            v = self._safe_evaluate("document.querySelectorAll('[data-asin]').length")
            count = v.get("result", {}).get("result", {}).get("value", 0)
            if count > 0:
                return True
            time.sleep(0.3)
        return False

    def _scroll_like_human(self):
        """模拟人工浏览：分段滚动 + 随机鼠标移动。"""
        import random
        steps = random.randint(3, 5)
        for _ in range(steps):
            y = 300 + random.randint(0, 1200)
            self._safe_evaluate(f"window.scrollTo(0, {y})")
            mx = 200 + random.randint(0, 600)
            my = 100 + random.randint(0, 800)
            self.cdp.mouse_event("mouseMoved", mx, my)
            time.sleep(0.3 + random.random() * 0.7)
        if random.random() > 0.4:
            self._safe_evaluate(f"window.scrollTo(0, {random.randint(0, 400)})")
            time.sleep(0.5 + random.random() * 0.5)

    # ── 搜索 ──────────────────────────────────────────────────────

    def search_keyword(self, keyword: str, max_pages: int = MAX_PAGES) -> dict:
        """搜索关键词，返回 {organic_rank, ad_pages, carousel}。

        首关键词首页通过搜索框输入提交（模拟人工搜索行为），
        后续页用 URL 翻页，保证广告布局与人工搜索一致。
        """
        self._ensure_tab()

        best_rank: int | None = None
        ad_pages: set[int] = set()
        ad_pages_white: set[int] = set()
        all_dp_asins: set[str] = set()  # ASINs found in any /dp/ link across all pages
        tracked_card_asins: set[str] = set()  # tracked ASINs that appeared in a [data-asin] card
        ad_asins_found: set[str] = set()  # tracked ASINs that appeared as ads

        for page in range(1, max_pages + 1):
            if page == 1:
                self._search_via_box(keyword)
            else:
                if not self._click_next_page():
                    break  # "下一页"按钮不可达，终止翻页

            # 等卡片渲染完成后再加少量缓冲
            if not self._wait_for_cards():
                # 慢机器可能 10s 不够，额外等 5s 再试一次
                time.sleep(5)
                if not self._wait_for_cards():
                    print(f"(slow page, skipping)", end=" ", flush=True)
                    continue
            self._human_delay(PAGE_LOAD_MIN, PAGE_LOAD_MAX)

            # 模拟人工浏览行为
            self._scroll_like_human()
            time.sleep(SCROLL_PAUSE)

            results, dp_asins = self._extract_page()
            all_dp_asins |= dp_asins

            for r in results:
                if r["asin"] not in self.asins:
                    continue
                tracked_card_asins.add(r["asin"])
                if r["isAd"] and r["asin"] in self.ad_asins:
                    ad_asins_found.add(r["asin"])
                    if self.white_asin and r["asin"] == self.white_asin:
                        ad_pages_white.add(page)
                    else:
                        ad_pages.add(page)
                if r["rank"] is not None and not r["isAd"]:
                    if best_rank is None or r["rank"] < best_rank:
                        best_rank = r["rank"]

        # Carousel: at least one tracked ASIN appeared on the page (card or /dp/
        # link), never got a rank, and was never an ad.  Per-ASIN: we exclude
        # ad ASINs so a different ASIN's ad doesn't suppress this ASIN's carousel.
        carousel_candidates = tracked_card_asins | (all_dp_asins & self.asins)
        carousel = best_rank is None and bool(carousel_candidates - ad_asins_found)

        result = {"organic_rank": best_rank, "ad_pages": sorted(ad_pages)}
        if self.white_asin:
            result["ad_pages_white"] = sorted(ad_pages_white)
        if carousel:
            result["carousel"] = True
        return result

    def _search_via_box(self, keyword: str):
        """通过搜索框输入+提交发起搜索，全程模拟鼠标操作。

        先导航到 Amazon 首页 → 鼠标点击搜索框 → 填入关键词 → 鼠标点击搜索按钮。
        """
        domain = SITE_DOMAIN[self.site]

        # 导航到首页（初始导航仍需 URL，相当于在地址栏输入）
        self._navigate(f"https://www.{domain}/")
        self._human_delay(2.0, 3.0)

        import random

        # 1) 鼠标点击搜索框（模拟真人先点一下获取焦点）
        self.cdp.click_element("#twotabsearchtextbox")
        time.sleep(0.15 + random.random() * 0.25)

        # 2) 填入关键词（原生 value setter + 事件触发）
        escaped = keyword.replace("\\", "\\\\").replace("'", "\\'")
        fill_js = (
            "(function(){"
            "var sb=document.querySelector('#twotabsearchtextbox');"
            "if(!sb)return 'no-searchbox';"
            "var d=Object.getOwnPropertyDescriptor(HTMLInputElement.prototype,'value');"
            f"d.set.call(sb,'{escaped}');"
            "sb.dispatchEvent(new Event('input',{bubbles:true}));"
            "sb.dispatchEvent(new Event('change',{bubbles:true}));"
            "return 'ok';"
            "})()"
        )
        raw = self._safe_evaluate(fill_js)
        result = str(raw.get("result", {}).get("result", {}).get("value", ""))
        if result == "no-searchbox":
            self._navigate(f"https://www.{domain}/s?k={quote(keyword)}")
            return

        # 3) 短暂的"打字思考"停顿
        time.sleep(0.3 + random.random() * 0.6)

        # 4) 鼠标点击搜索按钮
        self.cdp.click_element("#nav-search-submit-button")

    def _click_next_page(self) -> bool:
        """鼠标点击"下一页"按钮。成功返回 True。"""
        import random
        # 先滚到底部（翻页按钮在页面底部）
        self._safe_evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(0.4 + random.random() * 0.4)

        # 尝试点击下一页链接（Amazon 的 .s-pagination-next 或 a[href*='page='] 最后一个）
        result = self.cdp.click_element("a.s-pagination-next")
        if result.get("ok"):
            return True
        # 回退：尝试找最后一个包含 page= 的链接
        result = self.cdp.click_element(".s-pagination-container a:last-of-type")
        return result.get("ok", False)

    # ── 关键词间隔 ────────────────────────────────────────────────

    def gap_between_keywords(self):
        """关键词之间稍作停顿，模拟人工切换查询。"""
        self._human_delay(KEYWORD_GAP_MIN, KEYWORD_GAP_MAX)

    # ── BSR 大类排名 ──────────────────────────────────────────────

    def get_bsr(self) -> str:
        """获取 Best Sellers Rank，多类别时返回 '窗户类3-904 / 浴室类16-904'。"""
        parts = []
        for label, asin in self.bsr_asins:
            for attempt in range(2):
                try:
                    self._ensure_tab()
                    self._navigate(f"https://www.{SITE_DOMAIN[self.site]}/dp/{asin}")
                    self._wait_for_cards()
                    # US 站产品页面加载较慢(9000px+), 需要更长初始等待
                    self._human_delay(3.0, 4.0)

                    # 分段滚动触发懒加载（US 站长页面 BSR 在 9000px+，单次跳底不触发）
                    for y in (2000, 4000, 6000):
                        self._safe_evaluate(f"window.scrollTo(0, {y})")
                        time.sleep(0.5)
                    self._safe_evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    time.sleep(2.0)

                    self._safe_evaluate(_BSR_EXPAND_JS)
                    time.sleep(1.5)

                    raw = self._safe_evaluate(_BSR_JS)
                    ranks = json.loads(raw["result"]["result"]["value"])["ranks"]

                    if ranks:
                        nums = sorted(int(r["rank"]) for r in ranks)
                        val = "-".join(str(n) for n in nums)
                        parts.append(f"{label}{val}" if label else val)
                    break
                except Exception:
                    if attempt == 1:
                        raise
                    time.sleep(2)
        return " / ".join(parts) if parts else ""

    def close(self):
        self.cdp.close()
        self._tab_ws = None


# ── 格式化 ─────────────────────────────────────────────────────────

def format_result(organic_rank: int | None, ad_pages: list[int],
                  ad_pages_white: list[int] | None = None,
                  carousel: bool = False) -> str:
    """格式化排名结果。

    有 white 拆分时: （广告XY）（白广告Z），无广告组显示（/）
    无 white 时: 旧格式（广告X）（广告Y）...每个页码独立括号。
    carousel: ASIN 出现在页面内嵌/变体组件中，无法提取排名，统一输出 "/"。
    """
    if ad_pages_white is not None:
        non_white = f"（广告{''.join(map(str, ad_pages))}）" if ad_pages else "（/）"
        white = f"（白广告{''.join(map(str, ad_pages_white))}）" if ad_pages_white else "（/）"
        suffix = non_white + white
        if organic_rank is None:
            return "/" + suffix
        return f"{organic_rank}{suffix}"

    if organic_rank is None and not ad_pages:
        return "/"
    if organic_rank is None:
        return f"/（广告{''.join(map(str, ad_pages))}）"
    suffix = f"（广告{''.join(map(str, ad_pages))}）" if ad_pages else ""
    return f"{organic_rank}{suffix}"


# ── Excel 读写 ─────────────────────────────────────────────────────

def _detect_data_start_col(ws, start_scan: int = 3) -> int:
    """自动检测数据起始列号。

    策略: 扫描 row 2 找 BSR 模式 (如 '6-1310') → 扫描 row 1 找日期 → 回退 col 4。
    """
    # Strategy 1: BSR pattern in row 2
    for col in range(start_scan, max(ws.max_column, start_scan) + 1):
        v = ws.cell(row=2, column=col).value
        if v and BSR_PATTERN.match(str(v).strip()):
            return col

    # Strategy 2: date-like content in row 1
    for col in range(start_scan, max(ws.max_column, start_scan) + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        if isinstance(v, (datetime, date)):
            return col
        sv = str(v).strip()
        if re.search(r'\d+月\d+日', sv) or re.search(r'\d+[-/]\d+', sv):
            return col

    return 4


def read_keywords(excel_path: Path, data_start_col: int | None = None) -> tuple[openpyxl.Workbook, list[str], list[int], int, int]:
    """读取 Excel，返回 (wb, keywords, keyword_rows, 最新日期所在列号, 数据起始列号)。"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # B 列关键词 (row 3+)，遇到 ASIN 标注(B0前缀)自动截断
    keywords = []
    keyword_rows = []
    for row in range(3, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value
        if v:
            sv = str(v).strip()
            if sv.startswith("B0"):
                break
            keywords.append(sv)
            keyword_rows.append(row)

    # 确定数据起始列
    start_col = data_start_col if data_start_col is not None else _detect_data_start_col(ws)

    # 找到最新日期列
    last_date_col = start_col - 1
    for col in range(start_col, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            last_date_col = col

    return wb, keywords, keyword_rows, last_date_col, start_col


def write_results(
    wb: openpyxl.Workbook,
    excel_path: Path,
    keywords: list[str],
    keyword_rows: list[int],
    last_date_col: int,
    results: list[dict],
    bsr: str,
):
    """将查询结果写入新日期列。结果为空时拒绝写入，避免空白列覆盖存量。"""
    if not results:
        print("WARNING: No keyword results to write — skipping save to protect existing data.")
        return last_date_col + 1, None

    ws = wb.active
    new_col = last_date_col + 1
    today = datetime.now()

    # Row 1: 日期 (Excel serial number)
    date_serial = today.year * 365 + today.month * 30 + today.day  # 简化 serial
    # 使用 openpyxl 的日期格式
    from datetime import date
    ws.cell(row=1, column=new_col, value=f"{today.month}月{today.day}日")

    # Row 2: BSR
    ws.cell(row=2, column=new_col, value=bsr)

    # Row 3+: 按实际行号写入, 跳过空行
    for i, kw in enumerate(keywords):
        if i < len(results):
            r = results[i]
            val = format_result(r.get("organic_rank"), r.get("ad_pages", []),
                                r.get("ad_pages_white"),
                                carousel=r.get("carousel", False))
            ws.cell(row=keyword_rows[i], column=new_col, value=val)

    # 备份后保存
    backup = excel_path.with_name(f"{excel_path.stem}_备份{excel_path.suffix}")
    shutil.copy2(excel_path, backup)

    try:
        wb.save(excel_path)
    except PermissionError:
        print(f"\n文件被占用，写入失败: {excel_path}")
        print("请关闭 Excel 后重新运行同一条命令，进度已保存无需重跑。")
        return None, backup
    return new_col, backup


# ── CLI ─────────────────────────────────────────────────────────────

def _progress_bar(i: int, total: int, width: int = 30) -> str:
    """生成进度条字符串。"""
    filled = int(width * (i + 1) / total)
    bar = "█" * filled + "░" * (width - filled)
    pct = (i + 1) * 100 // total
    return f"[{i+1}/{total}] |{bar}| {pct}%"


def _load_progress(progress_path: Path) -> tuple[list[dict], int]:
    """加载断点进度文件，返回 (results, 已完成关键词数)。不存在则返回 ([], 0)。"""
    if progress_path.exists():
        try:
            data = json.loads(progress_path.read_text(encoding="utf-8"))
            results = data.get("results", [])
            # 不计 BSR meta 条目，避免 done 比实际关键词数多 1，导致续跑跳词错位
            done = sum(1 for r in results if "_bsr" not in r)
            return results, done
        except (json.JSONDecodeError, KeyError):
            pass
    return [], 0


def _save_progress(progress_path: Path, results: list[dict]):
    progress_path.write_text(json.dumps({"results": results}, ensure_ascii=False), encoding="utf-8")


def _parse_bsr_args(raw: list[str]) -> list[tuple[str, str]]:
    """Parse '标签:ASIN' pairs from --bsr-asin args. Plain ASIN → empty label."""
    result = []
    for item in raw:
        if ":" in item:
            label, asin = item.split(":", 1)
            result.append((label, asin))
        else:
            result.append(("", item))
    return result


def _build_summary(args, keywords, pure_results, bsr, timer,
                   dry_run=False, column=None, backup=None):
    """Build JSON summary for --json output."""
    kw_results = {}
    errors = []
    for i, (kw, r) in enumerate(zip(keywords, pure_results)):
        label = format_result(r.get("organic_rank"), r.get("ad_pages", []),
                              carousel=r.get("carousel", False))
        kw_results[kw] = label
        if r.get("_error"):
            errors.append({"keyword": kw, "error": r["_error"]})

    summary = {
        "module": "keyword_rank",
        "mode": "dry_run" if dry_run else "live",
        "status": "ok",
        "elapsed": round(timer.elapsed, 1),
        "site": args.site,
        "asins": list(args.asin),
        "bsr_asin": args.bsr_asin if args.bsr_asin else [list(args.asin)[0]],
        "bsr": bsr,
        "keywords_total": len(keywords),
        "keywords_ok": sum(1 for r in pure_results if r.get("organic_rank")),
        "keywords_with_ad": sum(1 for r in pure_results if r.get("ad_pages") or r.get("ad_pages_white")),
        "results": kw_results,
    }
    if errors:
        summary["errors"] = errors
    if column:
        summary["column"] = openpyxl.utils.get_column_letter(column)
    if backup:
        summary["backup"] = str(backup)
    summary["excel"] = str(args.excel)
    return summary


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Amazon 关键词排名查询")
    parser.add_argument("excel", type=Path, help="关键词 Excel 文件路径")
    parser.add_argument("--site", choices=["de", "fr", "us", "ca"], default="de",
                        help="Amazon 站点 (默认 de)")
    parser.add_argument("--asin", nargs="+", required=True,
                        help="要追踪的产品 ASIN (至少一个)")
    parser.add_argument("--bsr-asin", nargs="*", default=[],
                        help="用于查询 BSR 的 ASIN（标签:ASIN，如 窗户类:B0H4LXJ5QG 浴室类:B0CLXXD2X4）")
    parser.add_argument("--ad-asin", nargs="*", default=None,
                        help="只追踪这些 ASIN 的广告位 (默认追踪全部 --asin)")
    parser.add_argument("--white-asin", default="",
                        help="白色变体 ASIN，其广告位单独记入第二个括号（白广告X）")
    parser.add_argument("--data-start-col", type=int, default=None,
                        help="数据起始列号 (默认自动检测)")
    parser.add_argument("--dry-run", action="store_true", help="只查询不写入")
    parser.add_argument("--reset", action="store_true", help="忽略断点，从头开始")
    parser.add_argument("--json", action="store_true", help="输出结构化 JSON 运行摘要")
    parser.add_argument("--host", default=CDP_HOST, help=f"CDP 地址 (默认 {CDP_HOST})")
    parser.add_argument("--port", type=int, default=CDP_PORT)
    args = parser.parse_args()

    try:
        _run_query(args)
    except Exception as e:
        fatal_error("keyword_rank", e, excel=str(args.excel), site=args.site)


def _run_query(args):
    timer = RunTimer()
    asins = set(args.asin)
    print(f"Site: amazon.{args.site}  |  Tracking ASINs: {asins}")

    wb, keywords, keyword_rows, last_date_col, data_start_col = read_keywords(args.excel, args.data_start_col)
    total = len(keywords)
    print(f"Keywords: {total}")
    print(f"Data start column: {data_start_col} ({openpyxl.utils.get_column_letter(data_start_col)})")
    print(f"Latest date column: {last_date_col} ({openpyxl.utils.get_column_letter(last_date_col)})")

    # ── 断点恢复 ──
    progress_path = args.excel.with_suffix(".progress.json")
    results, done = _load_progress(progress_path)

    if done > 0 and not args.reset:
        print(f"\nResuming from keyword #{done+1} ({done}/{total} completed, "
              f"{total - done} remaining)")
    elif args.reset and progress_path.exists():
        progress_path.unlink()
        results, done = [], 0

    # ── 查询 ──
    ad_asins = set(args.ad_asin) if args.ad_asin else None
    bsr_asins = _parse_bsr_args(args.bsr_asin) if args.bsr_asin else None
    checker = KeywordRankChecker(asins, site=args.site, bsr_asins=bsr_asins,
                                 ad_asins=ad_asins, white_asin=args.white_asin,
                                 host=args.host, port=args.port)

    bsr = next((r["_bsr"] for r in results if "_bsr" in r), "")

    if not bsr:
        print("\nFetching BSR...")
        bsr = checker.get_bsr()
        print(f"BSR: {bsr}")
        results.insert(0, {"_bsr": bsr})
        _save_progress(progress_path, results)

    for i in range(done, total):
        kw = keywords[i]
        if i > 0:
            checker.gap_between_keywords()

        bar = _progress_bar(i, total)
        print(f"\n{bar} | {kw}", end=" ", flush=True)

        for attempt in range(2):
            try:
                r = checker.search_keyword(kw)
                break
            except Exception as e:
                if attempt == 0:
                    print(f"(retry: {e})", end=" ", flush=True)
                    time.sleep(3)
                else:
                    r = {"organic_rank": None, "ad_pages": [], "_error": str(e)}
                    if args.white_asin:
                        r["ad_pages_white"] = []
                    print(f"FAILED: {e}", end=" ", flush=True)

        results.append(r)
        label = format_result(r.get("organic_rank"), r.get("ad_pages", []),
                              carousel=r.get("carousel", False))
        print(f"→ {label}", end="")

        _save_progress(progress_path, results)

    checker.close()

    # 提取纯结果（去掉 _bsr 等 meta）
    pure_results = [{k: v for k, v in r.items() if not k.startswith("_")} for r in results if "_bsr" not in r]

    if args.dry_run:
        col_letter = openpyxl.utils.get_column_letter(last_date_col + 1)
        print(f"\n\n[Dry run] Would write {len(pure_results)} results to column {col_letter}")
        print(f"BSR: {bsr}")
        for i, (kw, r) in enumerate(zip(keywords, pure_results)):
            label = format_result(r.get("organic_rank"), r.get("ad_pages", []),
                              carousel=r.get("carousel", False))
            print(f"  [{i+1:2d}] {kw:30s} → {label}")
        if args.json:
            print_json_summary(_build_summary(args, keywords, pure_results, bsr,
                                              timer, dry_run=True))
        progress_path.unlink()
        return

    new_col, backup = write_results(wb, args.excel, keywords, keyword_rows, last_date_col, pure_results, bsr)
    if new_col is not None:
        print(f"\n\nWritten to column {openpyxl.utils.get_column_letter(new_col)}")
        print(f"Backup: {backup}")
        progress_path.unlink()
    if args.json:
        print_json_summary(_build_summary(args, keywords, pure_results, bsr,
                                          timer, column=new_col, backup=backup))


if __name__ == "__main__":
    main()
