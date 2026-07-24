"""数据录入模块 - 解析半结构化文本并追加写入 Excel。

用户A每天将用户B发送的物流信息粘贴到 OpenClaw，
自动解析并按日期排序插入到指定 Excel 文件。

用法:
    python -m src.data_entry <excel_path> [--batch] [--us | --de]
    --batch   多条录入（空行分隔）
    --us      US规则：复制原产品行到各仓库，ZIP XML直写
    --de      DE规则：品名+箱数匹配回填，ZIP XML直写
"""

from __future__ import annotations

import io
import re
import shutil
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_ROW = 2

# OOXML namespace
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

# US 用 cellXfs 样式索引
STYLE_BORDER = "4"   # borderId=1, fillId=0 (边框,无填充)
STYLE_YELLOW = "6"   # borderId=1, fillId=3 (边框,黄底)


# ── 列模板 ──────────────────────────────────────────────────────

def ensure_template(ws) -> None:
    """一次性初始化列模板（幂等：已存在则跳过）。

    插入/重命名列后，列宽会随表头名一并迁移，公式引用也会修正。
    """
    from openpyxl.utils import get_column_letter

    def _find(hdr: str) -> int | None:
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=HEADER_ROW, column=c).value or "").strip() == hdr:
                return c
        return None

    # 保存变更前列宽：{表头: 宽度}
    _widths: dict[str, float] = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip()
        cd = ws.column_dimensions.get(get_column_letter(c))
        if cd and cd.width and h:
            _widths[h] = cd.width

    # 保存插入前合并单元格（排除 Row 1 分区标签）
    _saved_merges: list[dict] = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row == 1 and mc.max_row == 1:
            continue
        _saved_merges.append({
            "min_row": mc.min_row, "max_row": mc.max_row,
            "min_col": mc.min_col, "max_col": mc.max_col,
            "value": ws.cell(row=mc.min_row, column=mc.min_col).value,
        })

    # 记录插入操作 [{at, amount}, ...]
    inserts: list[dict] = []

    # 标准表头样式：等线 11pt 加粗白色 + 深灰底 + 细线边框
    _hdr_font = Font(name="等线", size=11, bold=True, color="FFFFFFFF")
    _hdr_fill = PatternFill(patternType="solid", fgColor="FF5A5A5A")
    _hdr_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _style_hdr(col: int) -> None:
        c = ws.cell(row=HEADER_ROW, column=col)
        c.font = _hdr_font
        c.fill = _hdr_fill
        c.border = _hdr_border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. 箱数量/箱内数量后插入箱规(长)/箱规(宽)/箱规(高)/重量
    base2 = _find("箱数量") or _find("箱内数量")
    if base2 and not _find("箱规(长)"):
        ws.insert_cols(base2 + 1, 4)
        for i, hdr in enumerate(["箱规(长)", "箱规(宽)", "箱规(高)", "重量"]):
            ws.cell(row=HEADER_ROW, column=base2 + 1 + i).value = hdr
            _style_hdr(base2 + 1 + i)
        inserts.append({"at": base2 + 1, "amount": 4})

    # 2. 重命名（兼容旧→新规范，新文件自动跳过）
    _rename(ws, "成品编码", "品名")
    _rename(ws, "箱数量", "箱内数量")
    _rename(ws, "发船时间", "发车、发船时间")
    _rename(ws, "配送时段", "时效")
    _rename(ws, "发车、发船后配送时段", "时效")
    _rename(ws, "预计发货时间", "仓库发货时间")
    _rename(ws, "发货渠道", "实际发货渠道")
    for old_h, new_h in [("箱数量", "箱内数量"), ("发船时间", "发车、发船时间"),
                          ("配送时段", "时效"), ("发车、发船后配送时段", "时效"),
                          ("预计发货时间", "仓库发货时间"), ("发货渠道", "实际发货渠道"),
                          ("成品编码", "品名")]:
        if old_h in _widths:
            _widths[new_h] = _widths.pop(old_h)

    # 列宽随表头迁移
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip()
        cl = get_column_letter(c)
        if h and h in _widths:
            ws.column_dimensions[cl].width = _widths[h]
        elif h and cl in ws.column_dimensions:
            del ws.column_dimensions[cl]
    for hdr in ["箱规(长)", "箱规(宽)", "箱规(高)", "重量"]:
        col = _find(hdr)
        if col:
            ws.column_dimensions[get_column_letter(col)].width = 10

    # 修正 insert_cols 导致错位的合并单元格
    if inserts and _saved_merges:
        _fix_merged_cells(ws, inserts, _saved_merges)
    _fix_section_merges(ws)

    # 修正公式引用
    if inserts:
        _fix_formulas(ws, inserts)


def _rename(ws, old: str, new: str) -> None:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=HEADER_ROW, column=c).value or "").strip() == old:
            ws.cell(row=HEADER_ROW, column=c).value = new
            return


def _fix_merged_cells(ws, inserts: list[dict], saved: list[dict]) -> None:
    """修正 insert_cols 后所有合并单元格的列引用。

    使用插入前保存的数据，按列位移重建。
    """
    inserts_sorted = sorted(inserts, key=lambda x: x["at"])
    inserts_orig = []
    for ins in inserts_sorted:
        at_orig = ins["at"]
        for prev in inserts_orig:
            if prev["at"] <= at_orig:
                at_orig -= prev["amount"]
        inserts_orig.append({"at": at_orig, "amount": ins["amount"]})
    inserts_orig.sort(key=lambda x: x["at"])

    def _shift_col(orig_col: int) -> int:
        col = orig_col
        for ins in inserts_orig:
            if orig_col >= ins["at"]:
                col += ins["amount"]
        return col

    # 清除旧合并（排除 Row 1）
    ws.merged_cells.ranges = {
        mc for mc in ws.merged_cells.ranges
        if mc.min_row == 1 and mc.max_row == 1
    }

    # 按新列位重建
    for s in saved:
        new_min = _shift_col(s["min_col"])
        new_max = _shift_col(s["max_col"])
        cl_min = get_column_letter(new_min)
        cl_max = get_column_letter(new_max)
        ws.merge_cells(f"{cl_min}{s['min_row']}:{cl_max}{s['max_row']}")
        if s["value"] is not None:
            ws.cell(row=s["min_row"], column=new_min).value = s["value"]


def _fix_section_merges(ws) -> None:
    """修正 Row 1 的分区合并单元格（运营填写 / 仓库填写）。

    运营填写：A列 ~ 发货公司列（含）
    仓库填写：预计发货时间列 ~ 最后一列
    """
    from openpyxl.utils import get_column_letter
    end_ops = None
    start_wh = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip()
        if h == "发货公司":
            end_ops = c
        if h in ("预计发货时间", "仓库发货时间") and start_wh is None:
            start_wh = c

    # 清理旧的 Row 1 合并单元格（直接操作 ranges 集合，避免 unmerge_cells 的 insert_cols 后遗症）
    ws.merged_cells.ranges = {
        mc for mc in ws.merged_cells.ranges
        if not (mc.min_row == 1 and mc.max_row == 1)
    }

    # 重建
    if end_ops:
        ops_end = get_column_letter(end_ops)
        ws.merge_cells(f"A1:{ops_end}1")
        ws.cell(row=1, column=1).value = "运营填写"
    if start_wh:
        wh_start = get_column_letter(start_wh)
        wh_end = get_column_letter(ws.max_column)
        ws.merge_cells(f"{wh_start}1:{wh_end}1")
        ws.cell(row=1, column=start_wh).value = "仓库填写"

    # 设置分区标签样式
    _section_font = Font(name="等线", size=11, bold=True, color="FFFFFFFF")
    _section_fill = PatternFill(patternType="solid", fgColor="FF5A5A5A")
    _section_align = Alignment(horizontal="center", vertical="center")
    for c in [1, start_wh]:
        if c:
            cell = ws.cell(row=1, column=c)
            cell.font = _section_font
            cell.fill = _section_fill
            cell.alignment = _section_align


def _fix_formulas(ws, inserts: list[dict]) -> None:
    """修正所有公式中的列引用，补偿 openpyxl insert_cols 不更新引用的缺陷。

    inserts: [{at, amount}, ...] — 按执行顺序记录的插入操作（at 为当时列号）。
    需要将所有 at 归一化到插入前的原始列号再计算累计位移。
    """
    import re

    # 归一化 at 到原始列号：后续插入的 at 受之前插入影响，要减回去
    inserts_orig = []
    for ins in inserts:
        at_orig = ins["at"]
        # 之前的插入如果在 at 左侧，会把 at 向右推
        for prev in inserts_orig:
            if prev["at"] <= at_orig:
                at_orig -= prev["amount"]
        inserts_orig.append({"at": at_orig, "amount": ins["amount"]})
    inserts_orig.sort(key=lambda x: x["at"])

    def _new_col(orig_col: int) -> int:
        col = orig_col
        for ins in inserts_orig:
            if orig_col >= ins["at"]:
                col += ins["amount"]
        return col

    from openpyxl.utils import get_column_letter, column_index_from_string
    # 列引用边界：左侧不能是字母/下划线，右侧不能是字母数字/下划线
    # 避免误伤 DISPIMG 等函数中嵌入的 GUID（如 "C20967A"）
    cell_ref = re.compile(r"(?<![A-Za-z_])(\$?[A-Z]{1,3})(\$?\d+)(?![A-Za-z0-9_])")

    for row_idx in range(3, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue

            def _fix(m):
                cl = m.group(1)
                rn = m.group(2)
                try:
                    old_c = column_index_from_string(cl)
                    new_c = _new_col(old_c)
                except ValueError:
                    return m.group(0)
                return (get_column_letter(new_c) + rn) if new_c != old_c else m.group(0)

            cell.value = cell_ref.sub(_fix, cell.value)


# ── 解析 ────────────────────────────────────────────────────────

# 标签名 → 内部 key（顺序不敏感）
_LABEL_MAP = {
    "货物名称": "product",
    "品名": "product",
    "渠道": "channel",
    "指定发货渠道": "channel_specified",
    "实际发货渠道": "channel",
    "数量": "quantity",
    "箱数": "quantity",
    "箱规": "box_spec",
    "货件编号": "fba_code",
    "货件号": "fba_code",
    "FBA 编号": "fba_code",
    "FBA编号": "fba_code",
    "配送地址": "address",
    "发车、发船时间": "ship_date_text",
    "发车。发船后提取时间": "delivery_text",
    "发车、发船后提取时间": "delivery_text",
    "时效": "delivery_text",
    "单价价格": "price",
    "价格": "price",
    "有无附加（多少）": "surcharge",
    "仓库": "warehouse",
    "重量": "weight",
    "发货公司": "company",
    "sku": "sku_count",
    "SKU": "sku_count",
    "是否自搬货": "self_move",
    "自搬货": "self_move",
    "发货数量": "ship_qty",
    "备注": "remark",
    "仓库发货时间": "wh_ship_time",
}

# 长标签优先，避免子串误匹配（如 "渠道" ⊂ "指定发货渠道"）
_LABEL_MAP_SORTED = sorted(_LABEL_MAP.items(), key=lambda x: -len(x[0]))


def parse_entry(text: str) -> dict | None:
    """解析一段物流信息文本，返回结构化字段 dict。"""
    lines = [ln.strip() for ln in text.strip().split("\n") if ln.strip()]
    if len(lines) < 3:
        return None

    entry: dict = {}

    # 第 1 行：日期
    entry["date"] = _parse_date_line(lines[0])

    # 第 2 行：公司名或店铺名（带 - 国家后缀 → 发货店铺）
    line2 = lines[1].strip()
    if re.match(r"^[一-鿿]+-[A-Z]{2}$", line2):
        entry["store"] = line2
    else:
        entry["company"] = line2

    # 余下行：标签键值对（长标签优先，避免子串误匹配如 渠道 ⊂ 指定发货渠道）
    for ln in lines[2:]:
        for label, key in _LABEL_MAP_SORTED:
            m = re.match(re.escape(label) + r"\s*[:：]", ln)
            if m:
                val = ln[m.end():].strip()
                if val:
                    entry[key] = val
                break

    if "date" not in entry:
        return None
    return entry


def _parse_date_line(text: str) -> datetime:
    """'7月15日' → datetime(2026, 7, 15); '7.15' / '7-15' also supported."""
    m = re.match(r"(\d{1,2})\s*月\s*(\d{1,2})\s*日?", text)
    if m:
        return datetime(2026, int(m.group(1)), int(m.group(2)))
    # M.D 或 M-D 格式（无前导零）
    m2 = re.match(r"(\d{1,2})\s*[.\-]\s*(\d{1,2})$", text)
    if m2:
        return datetime(2026, int(m2.group(1)), int(m2.group(2)))
    # 备选完整格式
    for fmt in ["%Y-%m-%d", "%m/%d", "%m月%d日"]:
        try:
            dt = datetime.strptime(text, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=2026)
            return dt
        except ValueError:
            continue
    today = datetime.now()
    return datetime(today.year, today.month, today.day)


def parse_batch(text: str) -> list[dict]:
    """解析多条物流信息，以空行或 '---' 分隔。"""
    blocks = re.split(r"\n\s*\n|\n---\n?", text.strip())
    entries = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        entry = parse_entry(block)
        if entry:
            entries.append(entry)
    return entries


# ── 格式化 ──────────────────────────────────────────────────────

def _fmt_ship_date(text: str) -> str:
    """'7-23号左右' → '7月23日'"""
    text = text.strip()
    m = re.match(r"(\d{1,2})\s*[-–]\s*(\d{1,2})", text)
    if m:
        return f"{int(m.group(1))}月{int(m.group(2))}日"
    return text


def _fmt_delivery(text: str) -> str:
    """'40-50派送' → '40-50自然日'; '开船到签收38-52天' → '38-52自然日'"""
    text = text.strip()
    m = re.search(r"(\d+\s*[-–]\s*\d+)", text)
    if m:
        return f"{m.group(1).replace(' ', '')}自然日"
    return text


def _fmt_price(text: str) -> str:
    """'6.5' → '6.5'（原样返回）"""
    return text.strip()


def _fmt_surcharge(text: str) -> str:
    """'0' → '' (跳过), 非0照填"""
    text = text.strip()
    try:
        val = float(text)
        if val == 0:
            return ""
    except ValueError:
        pass
    return text


def _extract_warehouse(address: str) -> str:
    """提取仓库编号。
    'DTM2-Kaltband...' → 'DTM2'
    'Amazon - HAJ1 Zur Alten... (HAJ1)' → 'HAJ1'
    """
    address = address.strip()
    # 末尾括号标注 → 取括号内容
    m = re.search(r"\(([A-Z0-9]{2,6})\)", address)
    if m:
        return m.group(1)
    # 最前面的编号（字母数字开头，后跟 -）
    m = re.match(r"^([A-Z0-9]{2,6})-", address)
    if m:
        return m.group(1)
    # 直接取第一个词
    return address.split()[0].strip("-")


def _extract_box_spec(text: str) -> dict:
    """'59*44*55cm，重21kg' → {box_l: '59', box_w: '44', box_h: '55', weight: '21'}"""
    result = {"box_l": "", "box_w": "", "box_h": "", "weight": ""}
    # 尺寸
    dims = re.match(r"(\d+)\s*\*\s*(\d+)\s*\*\s*(\d+)", text)
    if dims:
        result["box_l"] = dims.group(1)
        result["box_w"] = dims.group(2)
        result["box_h"] = dims.group(3)
    # 重量
    w = re.search(r"重\s*(\d+(?:\.\d+)?)\s*kg", text, re.IGNORECASE)
    if w:
        result["weight"] = w.group(1)
    return result


# ── 写入 ────────────────────────────────────────────────────────

def _col_map(ws) -> dict[str, int]:
    """返回 {内部key: 列号}，基于当前表头。"""
    header_to_key = {
        "品名": "product",
        "货物名称": "product",
        "实际发货渠道": "channel",
        "指定发货渠道": "channel_specified",
        "发货渠道": "channel",
        "箱数": "quantity",
        "箱规(长)": "box_l",
        "箱规(宽)": "box_w",
        "箱规(高)": "box_h",
        "重量": "weight",
        "箱内数量": "box_inner_qty",
        "货件号": "fba_code",
        "仓库": "warehouse",
        "发货店铺": "store",
        "发货公司": "company",
        "发车、发船时间": "ship_date",
        "时效": "delivery",
        "发车、发船后配送时段": "delivery",
        "价格": "price",
        "附加费": "surcharge",
        "是否自搬货": "self_move",
        "发货数量": "ship_qty",
        "备注": "remark",
        "仓库发货时间": "wh_ship_time",
    }
    mapping = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip()
        if h in header_to_key:
            mapping[header_to_key[h]] = c
    return mapping


def _find_insert_row(ws, target_date: datetime) -> int:
    """在 A 列中找到 ≤ target_date 的最后一行，返回其后行号（即插入位）。

    空行跳过不中断，因为中间可能有空白分隔行。
    """
    last_match = HEADER_ROW
    for r in range(3, ws.max_row + 1):
        cell = ws.cell(row=r, column=1).value
        dt = None
        if isinstance(cell, datetime):
            dt = cell
        elif isinstance(cell, (int, float)):
            from datetime import timedelta
            dt = datetime(1899, 12, 30) + timedelta(days=int(cell))
        if dt is None:
            continue
        if dt.date() <= target_date.date():
            last_match = r
    return last_match + 1


def _resolve_date(entry: dict) -> datetime:
    """从 entry 提取日期，容错默认当天。"""
    d = entry.get("date")
    if isinstance(d, datetime):
        return d
    return datetime.now()


def _inherit_style(ws, ref_row: int, dst_row: int, col: int) -> None:
    """从参考行同列继承字体/边框/对齐/数字格式，不继承填充色。"""
    ref = ws.cell(row=ref_row, column=col)
    dst = ws.cell(row=dst_row, column=col)
    dst.font = copy(ref.font)
    dst.border = copy(ref.border)
    dst.alignment = copy(ref.alignment)
    if ref.number_format and ref.number_format != "General":
        dst.number_format = ref.number_format


def insert_batch(excel_path: str | Path, entries: list[dict]) -> dict:
    """批量插入多条解析后的条目，只打开/保存一次 Excel。"""
    excel_path = Path(excel_path)
    backup_path = excel_path.with_name(f"{excel_path.stem}_备份{excel_path.suffix}")
    shutil.copy2(excel_path, backup_path)

    wb = openpyxl.load_workbook(excel_path)
    sheet_name = _find_data_sheet(wb)
    if not sheet_name:
        wb.close()
        return {"error": "No data sheet found"}
    ws = wb[sheet_name]

    ensure_template(ws)
    cols = _col_map(ws)

    indexed = [(e, _resolve_date(e)) for e in entries]
    indexed.sort(key=lambda x: x[1])

    results = []
    for entry, entry_date in indexed:
        date_val = entry_date if isinstance(entry_date, datetime) else datetime.now()
        raw_data = _build_row_data(entry)
        insert_row = _find_insert_row(ws, date_val)

        ws.insert_rows(insert_row, 1)
        ref_row = insert_row - 1 if insert_row > 3 else insert_row + 1

        _copy_row_format(ws, ref_row, insert_row)
        for col in range(1, ws.max_column + 1):
            _inherit_style(ws, ref_row, insert_row, col)

        a_cell = ws.cell(row=insert_row, column=1)
        a_cell.value = date_val
        ref_a = ws.cell(row=ref_row, column=1)
        if ref_a.number_format and ref_a.number_format != "General":
            a_cell.number_format = ref_a.number_format
        else:
            a_cell.number_format = 'm"月"d"日";@'

        for key, col in cols.items():
            val = raw_data.get(key, "")
            if val:
                ws.cell(row=insert_row, column=col).value = val

        results.append({
            "row": insert_row,
            "date": date_val.strftime("%Y-%m-%d"),
        })

    try:
        wb.save(excel_path)
    except (PermissionError, OSError) as e:
        wb.close()
        return {"error": str(e), "backup": str(backup_path)}
    wb.close()

    return {
        "sheet": sheet_name,
        "inserted": len(results),
        "total_rows": len(results),
        "backup": str(backup_path),
    }


def insert_entry(excel_path: str | Path, entry: dict) -> dict:
    """将一条解析后的条目插入 Excel，返回结果。"""
    excel_path = Path(excel_path)
    backup_path = excel_path.with_name(f"{excel_path.stem}_备份{excel_path.suffix}")
    shutil.copy2(excel_path, backup_path)

    wb = openpyxl.load_workbook(excel_path)
    sheet_name = _find_data_sheet(wb)
    if not sheet_name:
        wb.close()
        return {"error": "No data sheet found"}
    ws = wb[sheet_name]

    ensure_template(ws)
    cols = _col_map(ws)

    entry_date = entry.get("date")
    date_val = entry_date if isinstance(entry_date, datetime) else datetime.now()
    raw_data = _build_row_data(entry)

    insert_row = _find_insert_row(ws, date_val)
    ws.insert_rows(insert_row, 1)
    ref_row = insert_row - 1 if insert_row > 3 else insert_row + 1

    _copy_row_format(ws, ref_row, insert_row)
    for col in range(1, ws.max_column + 1):
        _inherit_style(ws, ref_row, insert_row, col)

    a_cell = ws.cell(row=insert_row, column=1)
    a_cell.value = date_val
    ref_a = ws.cell(row=ref_row, column=1)
    if ref_a.number_format and ref_a.number_format != "General":
        a_cell.number_format = ref_a.number_format
    else:
        a_cell.number_format = 'm"月"d"日";@'

    filled = []
    for key, col in cols.items():
        val = raw_data.get(key, "")
        if not val:
            continue
        ws.cell(row=insert_row, column=col).value = val
        filled.append(f"{key}={val}")

    try:
        wb.save(excel_path)
    except (PermissionError, OSError) as e:
        wb.close()
        return {"error": str(e), "backup": str(backup_path)}
    wb.close()

    return {
        "sheet": sheet_name,
        "row": insert_row,
        "date": date_val.strftime("%Y-%m-%d"),
        "filled": filled,
        "backup": str(backup_path),
    }


def _build_row_data(entry: dict) -> dict:
    """从解析结果构建写入数据，含格式化转换。"""
    data = {}

    # 公司 / 店铺（互斥，二选一）
    data["company"] = entry.get("company", "")
    data["store"] = entry.get("store", "")

    # 货物名称
    data["product"] = entry.get("product", "")

    # 渠道（指定发货渠道 / 实际发货渠道）
    data["channel"] = entry.get("channel", "")
    data["channel_specified"] = entry.get("channel_specified", "")

    # 数量 → 提取数字
    qty = entry.get("quantity", "")
    m = re.search(r"(\d+)", str(qty))
    data["quantity"] = m.group(1) if m else qty

    # 箱规 → 拆成 3 个尺寸字段
    spec = _extract_box_spec(entry.get("box_spec", ""))
    data["box_l"] = spec["box_l"]
    data["box_w"] = spec["box_w"]
    data["box_h"] = spec["box_h"]

    # 重量：优先独立字段，其次箱规提取
    w = entry.get("weight", "")
    if w:
        data["weight"] = w.replace("KG", "").replace("kg", "").strip()
    elif spec["weight"]:
        data["weight"] = spec["weight"]

    # 货件编号
    data["fba_code"] = entry.get("fba_code", "")

    # 仓库：优先独立字段，其次地址提取
    wh = entry.get("warehouse", "")
    if wh:
        data["warehouse"] = wh
    else:
        addr = entry.get("address", "")
        data["warehouse"] = _extract_warehouse(addr) if addr else ""

    # 发车、发船时间
    ship = entry.get("ship_date_text", "")
    data["ship_date"] = _fmt_ship_date(ship) if ship else ""

    # 配送时段
    delivery = entry.get("delivery_text", "")
    data["delivery"] = _fmt_delivery(delivery) if delivery else ""

    # 价格
    price = entry.get("price", "")
    data["price"] = _fmt_price(price) if price else ""

    # 附加费
    surcharge = entry.get("surcharge", "")
    data["surcharge"] = _fmt_surcharge(surcharge) if surcharge else ""

    # 是否自搬货
    data["self_move"] = entry.get("self_move", "")

    # 发货数量
    data["ship_qty"] = entry.get("ship_qty", "")

    # 备注
    data["remark"] = entry.get("remark", "")

    # 仓库发货时间
    data["wh_ship_time"] = entry.get("wh_ship_time", "")

    return data


def _copy_row_format(ws, src_row: int, dst_row: int) -> None:
    """复制行格式（行高、字体对齐等）。"""
    src_dim = ws.row_dimensions.get(src_row)
    if src_dim:
        ws.row_dimensions[dst_row].height = src_dim.height


def _find_data_sheet(wb) -> str | None:
    """找到第一个以数字命名的 sheet。"""
    for sn in wb.sheetnames:
        if sn.isdigit() or any(c.isdigit() for c in sn):
            return sn
    return wb.sheetnames[0] if wb.sheetnames else None


# ── ZIP 工具 ──────────────────────────────────────────────────────

def _fix_wps_shared_strings(f: Path) -> None:
    """WPS 文件在 Content_Types/rels 中声明了 sharedStrings.xml 但未写入 ZIP，
    openpyxl 会因找不到该文件而崩溃。仅在文件确实缺失时移除引用。"""
    with zipfile.ZipFile(f) as zf:
        if "xl/sharedStrings.xml" in zf.namelist():
            return  # 文件有 sharedStrings.xml，无需修复

    tmp = f.with_suffix(".wpsfix.zip")
    with zipfile.ZipFile(f) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = re.sub(
                    r'<Override PartName="/xl/sharedStrings\.xml"[^>]*/>', "", text
                )
                data = text.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                text = re.sub(r'<Relationship[^>]*sharedStrings[^>]*/>', "", text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(f)


def _wps_restore_from_memory(target: Path, backup_data: dict[str, bytes]) -> None:
    """openpyxl save 后从内存备份恢复 WPS 专属内容：行高、cellImages、media。"""
    import re as _re
    restore: dict[str, bytes] = {}

    with zipfile.ZipFile(target) as zt:
        # ── media 文件 + cellimages XML ──
        for n, data in backup_data.items():
            if (n.startswith("xl/media/") and not n.endswith("/")) or \
               n in ("xl/cellimages.xml", "xl/_rels/cellimages.xml.rels"):
                restore[n] = data

        # ── 合并 Content_Types: 补回 image/cellImage 注册 ──
        if "[Content_Types].xml" in backup_data:
            b_ct = backup_data["[Content_Types].xml"].decode("utf-8")
            t_ct = zt.read("[Content_Types].xml").decode("utf-8")
            for full_tag in _re.findall(
                r"<Default[^>]*Extension=\"(?:png|jpe?g|gif|bmp)\"[^>]*/>", b_ct
            ):
                if full_tag not in t_ct:
                    t_ct = t_ct.replace("</Types>", full_tag + "\n</Types>")
            for full_tag in _re.findall(
                r"<Override[^>]*cellimage[^>]*/>", b_ct, _re.IGNORECASE
            ):
                if full_tag not in t_ct:
                    t_ct = t_ct.replace("</Types>", full_tag + "\n</Types>")
            restore["[Content_Types].xml"] = t_ct.encode("utf-8")

        # ── 合并 workbook.xml.rels: 补回 cellImage 关系 ──
        if "xl/_rels/workbook.xml.rels" in backup_data:
            b_rels = backup_data["xl/_rels/workbook.xml.rels"].decode("utf-8")
            t_rels = zt.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            for full_tag in _re.findall(
                r"<Relationship[^>]*cellImage[^>]*/>", b_rels, _re.IGNORECASE
            ):
                if full_tag not in t_rels:
                    t_rels = t_rels.replace("</Relationships>", full_tag + "\n</Relationships>")
            restore["xl/_rels/workbook.xml.rels"] = t_rels.encode("utf-8")

        # ── 合并 sheet XML: 补回行高 ──
        for sn, b_xml_bytes in backup_data.items():
            if not (sn.startswith("xl/worksheets/sheet") and sn.endswith(".xml")):
                continue
            b_xml = b_xml_bytes.decode("utf-8")
            t_xml = zt.read(sn).decode("utf-8")
            # 从备份提取行高
            b_heights = {}
            for row_tag in _re.findall(r"<row[^>]*>", b_xml):
                r = _re.search(r'r="(\d+)"', row_tag)
                ht = _re.search(r'ht="([^"]+)"', row_tag)
                ch = _re.search(r'customHeight="(\d+)"', row_tag)
                if r and ht:
                    ht_val = float(ht.group(1))
                    # ht=1 是 WPS 的自动行高标记，openpyxl 处理后不再生效，跳过让其走 defaultRowHeight
                    if ht_val <= 1:
                        continue
                    b_heights[int(r.group(1))] = (ht.group(1), ch.group(1) if ch else "0")

            def _patch_row(m):
                rn = int(m.group(1))
                attrs = m.group(2)
                if rn in b_heights:
                    ht_val, ch_val = b_heights[rn]
                    if 'ht="' in attrs:
                        attrs = _re.sub(r'ht="[^"]*"', f'ht="{ht_val}"', attrs)
                    else:
                        attrs += f' ht="{ht_val}" customHeight="{ch_val}"'
                return f'<row r="{rn}"{attrs}>'

            t_xml = _re.sub(r'<row r="(\d+)"([^>]*)>', _patch_row, t_xml)
            cur_rows = {int(m.group(1)) for m in _re.finditer(r'<row r="(\d+)"', t_xml)}
            for rn, (ht_val, ch_val) in b_heights.items():
                if rn not in cur_rows:
                    tag = f'<row r="{rn}" ht="{ht_val}" customHeight="{ch_val}"/>'
                    t_xml = t_xml.replace("</sheetData>", f"{tag}\n</sheetData>")
            restore[sn] = t_xml.encode("utf-8")

    if not restore:
        return

    tmp = target.with_suffix(".tmp.zip")
    with zipfile.ZipFile(target) as zt, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zt.infolist():
            if item.filename not in restore:
                zo.writestr(item, zt.read(item.filename))
        for name, data in restore.items():
            zo.writestr(name, data)
    tmp.replace(target)


# ── DE 解析 ───────────────────────────────────────────────────────

def parse_de_entry(text: str) -> dict | None:
    """解析 DE 格式物流文本。

    第1行：日期
    第2行：公司/店铺
    品名行：品名：1、产品名-箱数  2、产品名-箱数
    余下行：标签键值对
    """
    lines = [ln.strip() for ln in text.strip().split("\n") if ln.strip()]
    if len(lines) < 3:
        return None

    entry: dict = {"products": []}

    # 第1行：日期
    entry["date"] = _parse_date_line(lines[0])

    # 第2行：店铺 / 公司
    line2 = lines[1].strip()
    if re.match(r"^[一-鿿]+-[A-Z]{2}$", line2):
        entry["store"] = line2
    else:
        entry["company"] = line2

    # 品名行：品名：1、xxx-N箱  2、yyy-M箱
    prod_pattern = re.compile(r"(\d+)\s*[、,，.\-]\s*(.+)$")
    for ln in lines[2:]:
        cleaned = re.sub(r"^(品名|产品)\s*[:：]\s*", "", ln)
        matched_product = False
        for text in (cleaned, ln):
            m = prod_pattern.match(text)
            if m:
                name_part = m.group(2)
                # 末尾 "-N箱"
                qm = re.search(r"[-–](\d+)\s*箱\s*$", name_part)
                if qm:
                    entry["products"].append({
                        "name": name_part[:qm.start()].strip(),
                        "boxes": int(qm.group(1)),
                    })
                    matched_product = True
                break
        if matched_product:
            continue
        # 普通标签匹配
        for label, key in _LABEL_MAP_SORTED:
            m = re.match(re.escape(label) + r"\s*[:：]", ln)
            if m:
                val = ln[m.end():].strip()
                if val:
                    entry[key] = val
                break

    if not entry["products"]:
        return None
    return entry


# ── DE 写入 ───────────────────────────────────────────────────────

def _build_de_fields(entry: dict) -> dict[str, str]:
    """从 DE 解析结果构建要回填的字段，含格式化转换。"""
    result: dict[str, str] = {}

    # 简单透传
    for key in ("company", "store", "channel_specified", "channel", "fba_code",
                "warehouse", "self_move", "ship_qty", "remark", "wh_ship_time"):
        if entry.get(key):
            result[key] = entry[key]

    # 箱规 → 三个尺寸列
    spec = _extract_box_spec(entry.get("box_spec", ""))
    result["box_l"] = spec["box_l"]
    result["box_w"] = spec["box_w"]
    result["box_h"] = spec["box_h"]
    if not result.get("weight") and spec["weight"]:
        result["weight"] = spec["weight"]

    # 重量
    w = entry.get("weight", "")
    result["weight"] = w.replace("KG", "").replace("kg", "").strip() if w else result.get("weight", "")

    # 发车时间
    st = entry.get("ship_date_text", "")
    result["ship_date"] = _fmt_ship_date(st) if st else ""

    # 时效
    dt = entry.get("delivery_text", "")
    result["delivery"] = _fmt_delivery(dt) if dt else ""

    # 价格
    p = entry.get("price", "")
    result["price"] = p.strip() if p else ""

    # 附加费
    result["surcharge"] = _fmt_surcharge(entry.get("surcharge", ""))

    return result


def _col_letter(n: int) -> str:
    """1 → 'A', 27 → 'AA'."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _xml_escape(text: str) -> str:
    """转义 XML 特殊字符。"""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") \
               .replace('"', "&quot;").replace("'", "&apos;")


def insert_de(excel_path: str | Path, entry: dict) -> dict:
    """DE 回填：ZIP 级 XML 编辑，只改写单元格值，不动任何其他内容。

    不经过 openpyxl — 行高、cellImages、样式、sharedStrings 全部原封保留。
    """
    import re as _re
    excel_path = Path(excel_path)

    fields = _build_de_fields(entry)
    # 内部 key → 格式化后的值（非空）
    write_vals = {k: v for k, v in fields.items() if v}

    # ── 1. 读 ZIP，定位 sheet ──
    with zipfile.ZipFile(excel_path, "r") as zf:
        zip_data = {n: zf.read(n) for n in zf.namelist() if not n.endswith("/")}

        # 从 workbook.xml 找第一个 sheet 的文件名
        wb_xml = zip_data.get("xl/workbook.xml", b"").decode("utf-8")
        sheet_file = "xl/worksheets/sheet1.xml"
        m = _re.search(r'<sheet[^>]*name="([^"]*)"[^>]*sheetId', wb_xml)
        if m:
            for sn in _re.findall(r'<sheet[^>]*/>', wb_xml):
                nm = _re.search(r'name="([^"]*)"', sn)
                fid = _re.search(r'r:id="([^"]*)"', sn)
                if fid:
                    rid = fid.group(1)
                    rid_m = _re.search(
                        rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]*)"',
                        zip_data.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8")
                    )
                    if rid_m:
                        sheet_file = "xl/" + rid_m.group(1)
                        break

        sheet_xml = zip_data[sheet_file].decode("utf-8")

    # ── 2. 解析 Row 2 表头 → {header: col_letter} ──
    row2_cells = {}
    row2_m = _re.search(r'<row[^>]*r="2"[^>]*>(.*?)</row>', sheet_xml, _re.DOTALL)
    if not row2_m:
        return {"error": "Row 2 (header) not found"}
    for cell_m in _re.finditer(r'<c[^>]*r="([A-Z]+)2"[^>]*>(.*?)</c>', row2_m.group(1), _re.DOTALL):
        col_letter = cell_m.group(1)
        cell_content = cell_m.group(2)
        val = ""
        # inline string
        t_m = _re.search(r'<t[^>]*>(.*?)</t>', cell_content)
        if t_m:
            val = t_m.group(1)
        # shared string index — we can't resolve without sharedStrings, skip
        if not val:
            v_m = _re.search(r'<v>(\d+)</v>', cell_content)
            if v_m and "t=\"s\"" in cell_m.group(0):
                val = _resolve_shared_string(
                    zip_data.get("xl/sharedStrings.xml", b"").decode("utf-8"), int(v_m.group(1))
                )
        if val:
            row2_cells[val.strip()] = col_letter

    # 表头文字 → 内部 key
    _h2k = {
        "品名": "product", "货物名称": "product",
        "实际发货渠道": "channel", "指定发货渠道": "channel_specified", "发货渠道": "channel",
        "箱数": "quantity", "箱规(长)": "box_l", "箱规(宽)": "box_w", "箱规(高)": "box_h",
        "重量": "weight", "箱内数量": "box_inner_qty", "货件号": "fba_code",
        "仓库": "warehouse", "发货店铺": "store", "发货公司": "company",
        "发车、发船时间": "ship_date", "时效": "delivery", "价格": "price",
        "附加费": "surcharge", "是否自搬货": "self_move",
        "发货数量": "ship_qty", "备注": "remark", "仓库发货时间": "wh_ship_time",
    }
    key_to_col: dict[str, str] = {}
    for hdr, cl in row2_cells.items():
        if hdr in _h2k:
            key_to_col[_h2k[hdr]] = cl

    name_col = key_to_col.get("product")
    box_col = key_to_col.get("quantity")
    if not name_col or not box_col:
        return {"error": "Missing 品名 or 箱数 column in header"}

    # ── 3. 解析数据行 → 匹配产品 ──
    ss_xml = zip_data.get("xl/sharedStrings.xml", b"").decode("utf-8")

    row_tags = list(_re.finditer(
        r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', sheet_xml, _re.DOTALL
    ))

    # 为每行建立 {col_letter: cell_text}
    def _parse_row(row_xml: str) -> dict[str, str]:
        cells = {}
        for cm in _re.finditer(r'<c[^>]*r="([A-Z]+)(\d+)"[^>]*>(.*?)</c>', row_xml, _re.DOTALL):
            cl = cm.group(1)
            content = cm.group(3)
            full_tag = cm.group(0)
            val = ""
            t_m = _re.search(r'<t[^>]*>(.*?)</t>', content)
            if t_m:
                val = t_m.group(1)
            else:
                v_m = _re.search(r'<v>(.*?)</v>', content)
                if v_m:
                    if 't="s"' in full_tag:
                        val = _resolve_shared_string(ss_xml, int(v_m.group(1)))
                    else:
                        val = v_m.group(1)
            cells[cl] = val
        return cells

    results = []
    matched_rows: dict[int, dict[str, str]] = {}  # {row_num: {col_letter: value}}

    for row_m in row_tags:
        rn = int(row_m.group(1))
        if rn < 3:
            continue
        cells = _parse_row(row_m.group(2))
        cn = cells.get(name_col, "").strip()
        cb_str = cells.get(box_col, "")
        try:
            cb = int(float(cb_str)) if cb_str else None
        except (ValueError, TypeError):
            cb = None

        for prod in entry["products"]:
            if cn == prod["name"] and cb == prod["boxes"]:
                results.append({"name": prod["name"][:30], "row": rn, "filled": len(write_vals)})
                matched_rows[rn] = cells
                break

    # ── 4. 构建新 cell XML，写入匹配行 ──

    def _build_cell(col_letter: str, row_num: int, value: str) -> str:
        ref = f"{col_letter}{row_num}"
        return (
            f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">'
            f'{_xml_escape(value)}</t></is></c>'
        )

    # Build modified sheet XML row by row
    def _replace_row(m):
        rn = int(m.group(1))
        attrs = m.group(2)
        if rn not in matched_rows:
            return m.group(0)
        # Remove existing cells that we'll overwrite
        row_content = m.group(3)
        for key in write_vals:
            cl = key_to_col.get(key)
            if cl:
                ref = f'{cl}{rn}'
                row_content = _re.sub(
                    rf'<c[^>]*r="{ref}"[^>]*>(?:.*?</c>)',
                    '', row_content, flags=_re.DOTALL
                )
                row_content = _re.sub(
                    rf'<c[^>]*r="{ref}"[^>]*/>',
                    '', row_content
                )
        # Build new cells
        new_cells = ""
        for key, val in write_vals.items():
            cl = key_to_col.get(key)
            if cl:
                new_cells += _build_cell(cl, rn, val)
        return f'<row r="{rn}"{attrs}>{row_content}{new_cells}</row>'

    sheet_xml = _re.sub(
        r'<row[^>]*r="(\d+)"([^>]*)>(.*?)</row>',
        _replace_row, sheet_xml, flags=_re.DOTALL
    )

    # For rows not in the original XML but matched, append after </sheetData>
    # (this shouldn't happen for DE mode, but handle gracefully)

    # ── 5. 写回 ZIP ──
    zip_data[sheet_file] = sheet_xml.encode("utf-8")

    tmp = excel_path.with_suffix(".de_tmp.zip")
    with zipfile.ZipFile(excel_path, "r") as zr, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zw:
        for item in zr.infolist():
            name = item.filename
            if name in zip_data:
                zw.writestr(item, zip_data[name])
            else:
                zw.writestr(item, zr.read(name))
    tmp.replace(excel_path)

    matched = sum(1 for r in results if r["row"])
    return {
        "sheet": sheet_file.split("/")[-1].replace(".xml", ""),
        "matched": matched,
        "total_products": len(entry["products"]),
        "results": results,
    }


def _resolve_shared_string(ss_xml: str, idx: int) -> str:
    """从 sharedStrings.xml 解析第 idx 个字符串。"""
    import re as _re
    items = _re.findall(r'<si>.*?<t[^>]*>(.*?)</t>.*?</si>', ss_xml, _re.DOTALL)
    if 0 <= idx < len(items):
        return items[idx]
    return ""


# ── US 解析 ───────────────────────────────────────────────────────

def parse_us_batch(text: str) -> list[dict]:
    """解析 US 格式：共享头部 + 编号货件块。忽略输入中的"货物名称"，品名取原文件。"""
    text = text.strip()
    lines = text.split("\n")

    shared: dict[str, str] = {}
    first_item_idx = 0
    for i, ln in enumerate(lines):
        ln = ln.strip()
        if not ln:
            continue
        if re.match(r"\d+、", ln):
            first_item_idx = i
            break
        m = re.match(r"([^：:]+)[：:](.+)", ln)
        if m:
            key = m.group(1).strip()
            if key == "货物名称":
                continue
            shared[key] = m.group(2).strip()

    remaining = "\n".join(lines[first_item_idx:])
    item_blocks = re.split(r"\n(?=\d+、)", remaining)

    shipments: list[dict] = []
    for block in item_blocks:
        block = block.strip()
        if not block:
            continue
        block = re.sub(r"^\d+、", "", block).strip()
        item_lines = block.split("\n")

        item: dict[str, str] = {}
        segments = re.split(r"-+", item_lines[0].strip())
        for seg in segments:
            seg = seg.strip()
            if not seg:
                continue
            m = re.match(r"([^：:]+)[：:](.+)", seg)
            if m:
                k = m.group(1).strip()
                v = m.group(2).strip()
                if k in ("SKU", "sku"):
                    item["sku_count"] = v
                elif k == "货件号":
                    item["fba_code"] = v
                elif k == "仓库":
                    item["warehouse"] = v
                elif k == "箱数":
                    item["quantity_ref"] = v

        for ln in item_lines[1:]:
            ln = ln.strip()
            if not ln:
                continue
            m = re.match(r"([^：:]+)[：:](.+)", ln)
            if m:
                k = m.group(1).strip()
                v = m.group(2).strip()
                if k == "价格":
                    item["price"] = v
                elif k in ("发车、发船后配送时段", "时效"):
                    item["delivery_text"] = v
            else:
                item["channel"] = ln

        merged = dict(shared)
        if "发货公司" in merged:
            merged["company"] = merged.pop("发货公司")
        if "发货店铺" in merged:
            merged["store"] = merged.pop("发货店铺")
        if "指定发货渠道" in merged:
            merged["channel_specified"] = merged.pop("指定发货渠道")
        if "箱规" in merged:
            merged["box_spec"] = merged.pop("箱规")
        if "重量" in merged:
            merged["weight"] = merged.pop("重量")
        if "开船时间" in merged:
            merged["ship_date_text"] = merged.pop("开船时间")
        merged.update(item)
        shipments.append(merged)

    return shipments


# ── US 源产品提取 ─────────────────────────────────────────────────

def _parse_row2_headers(sheet_data, ss_texts: list[str]) -> dict[str, str]:
    """解析 Row 2 表头，返回 {header_text: col_letter}。支持 inlineStr 和 shared string。"""
    result: dict[str, str] = {}
    for row in sheet_data.findall(f"{{{NS}}}row"):
        if row.get("r") != "2":
            continue
        for c in row.findall(f"{{{NS}}}c"):
            ref = c.get("r", "")
            m = re.match(r"([A-Z]+)2", ref)
            if not m:
                continue
            cl = m.group(1)
            # inlineStr
            is_el = c.find(f"{{{NS}}}is")
            if is_el is not None:
                t_el = is_el.find(f"{{{NS}}}t")
                if t_el is not None and t_el.text:
                    result[t_el.text.strip()] = cl
                continue
            # shared string
            if c.get("t") == "s":
                v_el = c.find(f"{{{NS}}}v")
                if v_el is not None and v_el.text:
                    try:
                        idx = int(v_el.text)
                        if 0 <= idx < len(ss_texts):
                            result[ss_texts[idx].strip()] = cl
                    except ValueError:
                        pass
        return result
    return result


def extract_source_products(excel_path: Path, store: str | None = None) -> tuple[list[dict], list[int]]:
    """从目标 Excel 提取产品信息 + DISPIMG 公式。列位按 Row 2 表头动态解析。返回 (products, source_row_numbers)。"""
    with zipfile.ZipFile(excel_path) as zf:
        ss_texts: list[str] = []
        try:
            ss_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in ss_root.findall(f"{{{NS}}}si"):
                t = si.find(f"{{{NS}}}t")
                ss_texts.append(t.text if t is not None and t.text else "")
        except Exception:
            pass

        def _resolve(v: str | None) -> str:
            if v is None:
                return ""
            try:
                idx = int(v)
                return ss_texts[idx] if 0 <= idx < len(ss_texts) else v
            except ValueError:
                return v

        sroot = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
        sheet_data = sroot.find(f"{{{NS}}}sheetData")
        if sheet_data is None:
            return [], []

        header_to_col = _parse_row2_headers(sheet_data, ss_texts)
        img_col = header_to_col.get("图片", "B")
        date_col = header_to_col.get("发表日期", "A")
        product_col = header_to_col.get("品名", "C")
        asin_col = header_to_col.get("asin", "D")
        sku_col = header_to_col.get("sku", "E")
        fnsku_col = header_to_col.get("fnsku", "F")
        store_col = header_to_col.get("发货店铺", "G")

        products: list[dict] = []
        source_rows: list[int] = []
        for row in sheet_data.findall(f"{{{NS}}}row"):
            rn = int(row.get("r", "0"))
            if rn < 3:
                continue

            cells: dict[str, str] = {}
            dispimg_f: str | None = None
            dispimg_v: str | None = None

            for c in row.findall(f"{{{NS}}}c"):
                ref = c.get("r", "")
                col = re.match(r"([A-Z]+)", ref).group(1) if ref else ""
                t = c.get("t")
                v_el = c.find(f"{{{NS}}}v")
                f_el = c.find(f"{{{NS}}}f")
                is_el = c.find(f"{{{NS}}}is")

                if col == img_col:
                    if f_el is not None and f_el.text:
                        dispimg_f = f_el.text
                    if v_el is not None and v_el.text:
                        dispimg_v = v_el.text
                elif v_el is not None and v_el.text:
                    cells[col] = _resolve(v_el.text) if t == "s" else v_el.text
                elif is_el is not None:
                    t_el = is_el.find(f"{{{NS}}}t")
                    if t_el is not None:
                        cells[col] = t_el.text or ""

            if not cells.get(product_col) or not cells.get(asin_col):
                continue
            if store and store_col and cells.get(store_col, "") != store:
                continue

            products.append({
                "product": cells.get(product_col, ""),
                "asin": cells.get(asin_col, ""),
                "sku": cells.get(sku_col, ""),
                "fnsku": cells.get(fnsku_col, ""),
                "date": cells.get(date_col, ""),
                "dispimg_f": dispimg_f or "",
                "dispimg_v": dispimg_v or "",
            })
            source_rows.append(rn)

        return products, source_rows


# ── US 写入 (ZIP XML) ──────────────────────────────────────────────

# header → internal field for US column resolution
_US_FIELD_HEADERS = {
    "date": "发表日期",
    "image": "图片",
    "product": "品名",
    "asin": "asin",
    "sku": "sku",
    "fnsku": "fnsku",
    "store": "发货店铺",
    "channel_specified": "指定发货渠道",
    "box_l": "箱规(长)",
    "box_w": "箱规(宽)",
    "box_h": "箱规(高)",
    "weight": "重量",
    "channel": "实际发货渠道",
    "company": "发货公司",
    "ship_date": "发车、发船时间",
    "delivery": "时效",
    "price": "价格",
    "warehouse": "仓库",
    "fba_code": "货件号",
}


def _us_parse_headers(sroot: ET.Element,
                      ss_dict: dict[int, str] | None = None) -> tuple[dict[str, str], list[str]]:
    """从 sheet XML 的 Row 2 解析表头，返回 ({field: col_letter}, all_col_letters)。"""
    sheet_data = sroot.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        return {}, []

    row2 = None
    for row in sheet_data.findall(f"{{{NS}}}row"):
        if row.get("r") == "2":
            row2 = row
            break
    if row2 is None:
        return {}, []

    col_to_header: dict[str, str] = {}
    all_cols: list[str] = []
    for cell in row2.findall(f"{{{NS}}}c"):
        ref = cell.get("r", "")
        m = re.match(r"([A-Z]+)2", ref)
        if not m:
            continue
        cl = m.group(1)
        all_cols.append(cl)
        # inlineStr
        is_el = cell.find(f"{{{NS}}}is")
        if is_el is not None:
            t_el = is_el.find(f"{{{NS}}}t")
            if t_el is not None and t_el.text:
                col_to_header[cl] = t_el.text.strip()
                continue
        # shared string (t="s")
        if cell.get("t") == "s" and ss_dict:
            v_el = cell.find(f"{{{NS}}}v")
            if v_el is not None and v_el.text:
                idx = int(v_el.text)
                text = ss_dict.get(idx, "")
                if text:
                    col_to_header[cl] = text.strip()

    header_to_col = {h: cl for cl, h in col_to_header.items()}
    field_to_col: dict[str, str] = {}
    for field, header in _US_FIELD_HEADERS.items():
        cl = header_to_col.get(header, "")
        if cl:
            field_to_col[field] = cl

    return field_to_col, sorted(all_cols, key=_col_sort_key)


def _col_sort_key(cl: str) -> int:
    """'A'→0, 'Z'→25, 'AA'→26, ..."""
    n = 0
    for ch in cl:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _us_build_cells(shipment: dict, product: dict,
                    col_map: dict[str, str]) -> list[tuple]:
    """构建一行的 (col, value, style, cell_type) 列表。列字母由表头动态解析。"""
    cells: list[tuple] = []
    ship = shipment.get("ship_date_text", "")

    c = col_map.get("date", "")
    date_val = product.get("date", "")
    if date_val and c:
        cells.append((c, date_val, STYLE_BORDER, "inlineStr"))

    dispimg_f = product.get("dispimg_f", "")
    c = col_map.get("image", "")
    if dispimg_f and c:
        cells.append((c, dispimg_f, STYLE_BORDER, "formula_str"))

    c = col_map.get("product", "")
    if product.get("product") and c:
        cells.append((c, product["product"], STYLE_YELLOW, "inlineStr"))
    c = col_map.get("asin", "")
    if product.get("asin") and c:
        cells.append((c, product["asin"], STYLE_BORDER, "inlineStr"))
    c = col_map.get("sku", "")
    if product.get("sku") and c:
        cells.append((c, product["sku"], STYLE_BORDER, "inlineStr"))
    c = col_map.get("fnsku", "")
    if product.get("fnsku") and c:
        cells.append((c, product["fnsku"], STYLE_BORDER, "inlineStr"))

    c = col_map.get("store", "")
    store = shipment.get("store", "")
    if store and c:
        cells.append((c, store, STYLE_BORDER, "inlineStr"))

    c = col_map.get("channel_specified", "")
    ch_spec = shipment.get("channel_specified", "")
    if ch_spec and c:
        cells.append((c, ch_spec, STYLE_BORDER, "inlineStr"))

    box_spec = shipment.get("box_spec", "")
    if box_spec:
        spec = _extract_box_spec(box_spec)
        cl = col_map.get("box_l", "")
        if spec["box_l"] and cl:
            cells.append((cl, spec["box_l"], STYLE_BORDER, "inlineStr"))
        cl = col_map.get("box_w", "")
        if spec["box_w"] and cl:
            cells.append((cl, spec["box_w"], STYLE_BORDER, "inlineStr"))
        cl = col_map.get("box_h", "")
        if spec["box_h"] and cl:
            cells.append((cl, spec["box_h"], STYLE_BORDER, "inlineStr"))

    c = col_map.get("weight", "")
    weight = shipment.get("weight", "")
    if weight and c:
        cells.append((c, weight, STYLE_BORDER, "inlineStr"))

    c = col_map.get("channel", "")
    channel = shipment.get("channel", "")
    if channel and c:
        cells.append((c, channel, STYLE_BORDER, "inlineStr"))

    c = col_map.get("company", "")
    company = shipment.get("company", "")
    if company and c:
        cells.append((c, company, STYLE_BORDER, "inlineStr"))

    c = col_map.get("ship_date", "")
    if ship and c:
        cells.append((c, _fmt_ship_date(ship), STYLE_BORDER, "inlineStr"))

    c = col_map.get("delivery", "")
    delivery = shipment.get("delivery_text", "")
    if delivery and c:
        cells.append((c, _fmt_delivery(delivery), STYLE_BORDER, "inlineStr"))

    c = col_map.get("price", "")
    price = shipment.get("price", "")
    if price and c:
        cells.append((c, price, STYLE_BORDER, "inlineStr"))

    c = col_map.get("warehouse", "")
    warehouse = shipment.get("warehouse", "")
    if warehouse and c:
        cells.append((c, warehouse, STYLE_BORDER, "inlineStr"))

    c = col_map.get("fba_code", "")
    fba = shipment.get("fba_code", "")
    if fba and c:
        cells.append((c, fba, STYLE_BORDER, "inlineStr"))

    return cells


def _us_build_row(rn: int, cells: list[tuple], all_cols: list[str]) -> ET.Element:
    """创建 <row> 元素，数据格 + 空边框格填满所有表头列。"""
    row = ET.Element(f"{{{NS}}}row")
    row.set("r", str(rn))
    row.set("ht", "100")
    row.set("customHeight", "1")

    filled_cols = set()
    for col_letter, value, style, cell_type in cells:
        filled_cols.add(col_letter)
        c = ET.SubElement(row, f"{{{NS}}}c")
        c.set("r", f"{col_letter}{rn}")
        if style:
            c.set("s", style)

        if cell_type == "inlineStr":
            c.set("t", "inlineStr")
            is_el = ET.SubElement(c, f"{{{NS}}}is")
            t_el = ET.SubElement(is_el, f"{{{NS}}}t")
            t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            t_el.text = str(value)
        elif cell_type == "formula_str":
            c.set("t", "str")
            f_el = ET.SubElement(c, f"{{{NS}}}f")
            f_el.text = str(value)
            dispimg_v = "=" + str(value).replace("_xlfn.", "")
            v_el = ET.SubElement(c, f"{{{NS}}}v")
            v_el.text = dispimg_v

    # 空边框格
    for col_letter in all_cols:
        if col_letter not in filled_cols:
            c = ET.SubElement(row, f"{{{NS}}}c")
            c.set("r", f"{col_letter}{rn}")
            c.set("s", STYLE_BORDER)

    return row


def insert_us(excel_path: str | Path, shipments: list[dict],
              products: list[dict], source_rows: list[int]) -> dict:
    """US 录入：ZIP XML 写入新行 + 删除原产品行。

    1. 备份
    2. 解析 sheet1.xml → 删除 source_rows
    3. 移位剩余行 ≥ max(source_rows)+1
    4. 插入新行（全列边框）
    5. 排序 → 写回 ZIP
    """
    excel_path = Path(excel_path)
    backup_path = excel_path.with_name(f"{excel_path.stem}_备份{excel_path.suffix}")
    shutil.copy2(excel_path, backup_path)

    with open(excel_path, "rb") as f:
        zip_bytes = f.read()

    num_new = len(shipments) * len(products)
    num_to_delete = len(source_rows)
    insert_at = min(source_rows)
    net_shift = num_new - num_to_delete

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        sheet_raw = zf.read("xl/worksheets/sheet1.xml")
        # Resolve shared strings for header parsing
        ss_dict: dict[int, str] = {}
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_xml = zf.read("xl/sharedStrings.xml").decode("utf-8")
            ss_dict = {i: t for i, t in enumerate(
                re.findall(r"<si>.*?<t[^>]*>(.*?)</t>.*?</si>", ss_xml, re.DOTALL)
            )}

    ET.register_namespace("", NS)
    sroot = ET.fromstring(sheet_raw)
    sheet_data = sroot.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("No sheetData found in sheet XML")

    col_map, all_cols = _us_parse_headers(sroot, ss_dict)
    if not col_map:
        raise ValueError("Could not parse header row (Row 2) — missing or empty headers")

    # ── 删除原产品行 ──
    to_remove = []
    for row in sheet_data.findall(f"{{{NS}}}row"):
        if int(row.get("r", "0")) in source_rows:
            to_remove.append(row)
    for row in to_remove:
        sheet_data.remove(row)

    # ── 移位剩余行 ──
    max_deleted = max(source_rows)
    for row in sheet_data.findall(f"{{{NS}}}row"):
        rn = int(row.get("r", "0"))
        if rn > max_deleted:
            new_rn = rn + net_shift
            row.set("r", str(new_rn))
            for cell in row.findall(f"{{{NS}}}c"):
                ref = cell.get("r", "")
                m = re.match(r"([A-Z]+)(\d+)", ref)
                if m and int(m.group(2)) > max_deleted:
                    cell.set("r", f"{m.group(1)}{int(m.group(2)) + net_shift}")

    # ── 插入新行 ──
    rn = insert_at
    for shipment in shipments:
        for product in products:
            row_el = _us_build_row(rn, _us_build_cells(shipment, product, col_map), all_cols)
            sheet_data.append(row_el)
            rn += 1

    # ── 排序 ──
    all_rows = sheet_data.findall(f"{{{NS}}}row")
    all_rows.sort(key=lambda r: int(r.get("r", "0")))
    for row in all_rows:
        sheet_data.remove(row)
    for row in all_rows:
        sheet_data.append(row)

    sheet_new = ET.tostring(sroot, encoding="unicode")

    # ── 写回 ZIP ──
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zin:
        out_buf = io.BytesIO()
        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "xl/worksheets/sheet1.xml":
                    zout.writestr(item, sheet_new)
                else:
                    zout.writestr(item, zin.read(item.filename))
        out_bytes = out_buf.getvalue()

    with open(excel_path, "wb") as f:
        f.write(out_bytes)

    return {
        "sheet": "sheet1",
        "shipments": len(shipments),
        "products_per": len(products),
        "inserted_rows": num_new,
        "start_row": insert_at,
        "end_row": insert_at + num_new - 1,
        "backup": str(backup_path),
    }


# ── CLI ─────────────────────────────────────────────────────────

def main():
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m src.data_entry <excel_path> [--batch] [--us | --de]")
        print("       Reads text from stdin and writes to Excel.")
        print("       --batch   Parse multiple entries (separated by blank lines)")
        print("       --us      US rule: copy source products to warehouses, ZIP XML")
        print("       --de      DE rule: match existing rows by 品名+箱数, backfill")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    args = sys.argv[2:]
    batch_mode = "--batch" in args
    us_rule = "--us" in args
    de_rule = "--de" in args
    text = sys.stdin.read().strip()

    if not text:
        print("No input text provided.")
        sys.exit(1)

    # ── US 模式：复制原产品行到各仓库，ZIP XML 直写 ──
    if us_rule:
        shipments = parse_us_batch(text)
        if not shipments:
            print("Failed to parse US input text.")
            sys.exit(1)
        print(f"Parsed {len(shipments)} shipments.")

        products, source_rows = extract_source_products(excel_path)
        if not products:
            print("Error: No source product rows found in target Excel.")
            sys.exit(1)
        print(f"Found {len(products)} source products (rows {source_rows}).")

        result = insert_us(excel_path, shipments, products, source_rows)
        if "error" in result:
            print(f"Error: {result['error']}")
            sys.exit(1)
        print(f"Inserted {result['inserted_rows']} rows (rows {result['start_row']}-{result['end_row']}) "
              f"into Sheet={result['sheet']}, deleted {len(source_rows)} old product rows.")
        return

    # ── DE 模式：匹配已有行，回填 ──
    if de_rule:
        entry = parse_de_entry(text)
        if not entry:
            print("Failed to parse DE input text.")
            sys.exit(1)
        print(f"Parsed {len(entry['products'])} products.")
        result = insert_de(excel_path, entry)
        if "error" in result:
            print(f"Error: {result['error']}")
            sys.exit(1)
        for r in result["results"]:
            status = f"Row{r['row']}" if r["row"] else "NOT FOUND"
            print(f"  {r['name'][:40]}: {status}")
        print(f"Matched {result['matched']}/{result['total_products']} products "
              f"in Sheet={result['sheet']}")
        return

    # ── 批量模式 ──
    if batch_mode:
        entries = parse_batch(text)
        if not entries:
            print("Failed to parse any entries from input text.")
            sys.exit(1)
        print(f"Parsed {len(entries)} entries.")
        result = insert_batch(excel_path, entries)
        if "error" in result:
            print(f"Error: {result['error']}")
            sys.exit(1)
        print(f"Inserted {result['inserted']} groups / {result['total_rows']} rows "
              f"into Sheet={result['sheet']}")
        return

    # ── 单条模式 ──
    entry = parse_entry(text)
    if not entry:
        print("Failed to parse input text.")
        sys.exit(1)

    result = insert_entry(excel_path, entry)
    if "error" in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    print(f"Inserted at Sheet={result['sheet']} Row={result['row']} Date={result['date']}")
    print(f"Fields: {', '.join(result['filled'])}")


if __name__ == "__main__":
    main()
