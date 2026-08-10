"""跨表数据填写：根据发货信息表更新统计表的在采/在途数量。

业务场景:
  统计表记录每款产品的在采/稳再在途/柘流在途数量。
  货件发出后生成发货信息表，需将发货数量从统计表"在采"扣除，
  并根据发货店铺（稳再/柘流）累加到对应的"在途"列。

去重:
  发货表可反复提交，脚本通过日期列识别新增数据，已处理日期的行自动跳过。
  追踪记录保存在统计表同目录的 .cross_track.json 中，不影响 Excel。

用法:
  python -m src.cross_table <统计表路径> <发货表路径...>
"""

from __future__ import annotations

import json
import shutil
from datetime import date, datetime
from pathlib import Path

import openpyxl

try:
    from .cli_utils import fatal_error, print_json_summary, RunTimer
except ImportError:
    from cli_utils import fatal_error, print_json_summary, RunTimer

SHIPPING_HEADER_ROW = 2        # 发货信息表表头行
STATS_HEADER_ROW = 1           # 统计表标签行
STATS_DATA_START = 2           # 统计表数据起始行
ASIN_PATTERN_PREFIX = "B0"     # ASIN 值前缀
DATE_HEADERS = {"预计发货时间", "发货日期", "日期", "开船时间", "date", "ship_date"}


# ── 表头 / 列位匹配 ──

def _find_header_col(ws, name: str, header_row: int) -> int | None:
    """在指定行按表头文字精确匹配，返回列号(1-based)。"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip() == name:
            return c
    return None


def _find_asin_col(ws, data_start: int, sample: int = 5) -> int | None:
    """按值前缀匹配 ASIN 列——统计表无 asin 表头时回退到数据模式识别。"""
    for c in range(1, ws.max_column + 1):
        hits = 0
        for r in range(data_start, min(data_start + sample, ws.max_row + 1)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and str(v).strip().startswith(ASIN_PATTERN_PREFIX):
                hits += 1
        if hits >= 2:
            return c
    return None


# ── 日期解析 ──

def _parse_date(val) -> date | None:
    """将 Excel 单元格值解析为 date，失败返回 None。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    import re
    m = re.match(r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})[日]?", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"(\d{1,2})[月/-](\d{1,2})[日]?", s)
    if m:
        return date(date.today().year, int(m.group(1)), int(m.group(2)))
    return None


# ── 日期追踪文件 ──

def _load_tracker(tracker_path: Path) -> dict[str, date]:
    if tracker_path.exists():
        try:
            data = json.loads(tracker_path.read_text(encoding="utf-8"))
            return {k: date.fromisoformat(v) for k, v in data.items()}
        except (json.JSONDecodeError, ValueError):
            pass
    return {}


def _save_tracker(tracker_path: Path, tracker: dict[str, date]):
    tracker_path.write_text(
        json.dumps({k: v.isoformat() for k, v in tracker.items()},
                   ensure_ascii=False, indent=2),
        encoding="utf-8")


# ── 读取发货信息表 ──

def _read_shipping(path: Path, cutoff: date | None = None) -> tuple[list[dict], date | None]:
    """从发货信息表提取 [{asin, store, qty, src_row, src_sheet, date}]。

    cutoff: 只处理日期 > cutoff 的行（None 表示全部处理）。
    自动检测日期列（预计发货时间/发货日期/日期/开船时间）。

    Returns:
        (shipments, max_date_in_file) — 无日期列时 max_date 为 None。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    result: list[dict] = []
    max_date: date | None = None
    skipped = 0
    any_date_col = False

    for sn in wb.sheetnames:
        ws = wb[sn]
        asin_col = _find_header_col(ws, "asin", SHIPPING_HEADER_ROW)
        store_col = _find_header_col(ws, "发货店铺", SHIPPING_HEADER_ROW)
        qty_col = _find_header_col(ws, "发货数量", SHIPPING_HEADER_ROW)

        # 检测日期列
        date_col: int | None = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=SHIPPING_HEADER_ROW, column=c).value
            if v and str(v).strip() in DATE_HEADERS:
                date_col = c
                any_date_col = True
                break

        if asin_col is None:
            print(f"  [警告] 发货表 Sheet '{sn}' 未找到 'asin' 表头, 跳过")
            continue
        if store_col is None:
            print(f"  [警告] 发货表 Sheet '{sn}' 未找到 '发货店铺' 表头, 跳过")
            continue
        if qty_col is None:
            print(f"  [警告] 发货表 Sheet '{sn}' 未找到 '发货数量' 表头, 跳过")
            continue

        for r in range(SHIPPING_HEADER_ROW + 1, ws.max_row + 1):
            asin = _safe_str(ws.cell(row=r, column=asin_col).value)
            if not asin or not asin.startswith(ASIN_PATTERN_PREFIX):
                continue
            store = _safe_str(ws.cell(row=r, column=store_col).value)
            qty = _safe_int(ws.cell(row=r, column=qty_col).value)
            if not store or qty == 0:
                continue

            row_date = None
            if date_col is not None:
                row_date = _parse_date(ws.cell(row=r, column=date_col).value)

            # 日期过滤：已处理日期的行跳过
            if cutoff is not None and row_date is not None and row_date <= cutoff:
                skipped += 1
                continue

            if row_date is not None and (max_date is None or row_date > max_date):
                max_date = row_date

            result.append({
                "asin": asin,
                "store": store,
                "qty": qty,
                "src_row": r,
                "src_sheet": sn,
                "date": row_date,
            })

    wb.close()

    if skipped > 0:
        print(f"  已跳过 {skipped} 行 (日期 <= {cutoff}, 已处理)")
    if not any_date_col:
        print(f"  [提示] 未检测到日期列，本次全量处理（后续提交将无去重保护）")

    return result, max_date


# ── 应用更新到统计表 ──

def _apply_updates(stats_path: Path, shipments: list[dict]) -> dict:
    """对统计表逐行应用在采/在途更新，返回 {updated, missing, negative, backup}。"""
    wb = openpyxl.load_workbook(stats_path)

    # 统计表通常只有 Sheet1
    ws = wb.active
    if ws is None:
        wb.close()
        return {"updated": 0, "missing": len(shipments), "negative": 0}

    # 匹配统计表列位
    asin_col = _find_header_col(ws, "asin", STATS_HEADER_ROW)
    if asin_col is None:
        # 回退：在数据行中按 B0 前缀识别 ASIN 列
        asin_col = _find_asin_col(ws, STATS_DATA_START)

    zaicai_col = _find_header_col(ws, "在采", STATS_HEADER_ROW)
    wz_zt_col = _find_header_col(ws, "稳再在途", STATS_HEADER_ROW)
    zl_zt_col = _find_header_col(ws, "柘流在途", STATS_HEADER_ROW)

    if asin_col is None:
        print("[错误] 统计表未找到 ASIN 列 (无 'asin' 表头且数据模式匹配失败)")
        wb.close()
        return {"updated": 0, "missing": len(shipments), "negative": 0}
    if zaicai_col is None:
        print("[错误] 统计表未找到 '在采' 列")
        wb.close()
        return {"updated": 0, "missing": len(shipments), "negative": 0}

    # 构建 ASIN → row 索引
    asin_row: dict[str, int] = {}
    for r in range(STATS_DATA_START, ws.max_row + 1):
        a = _safe_str(ws.cell(row=r, column=asin_col).value)
        if a:
            asin_row[a] = r

    updated = 0
    missing: list[str] = []
    negative: list[str] = []

    for s in shipments:
        asin = s["asin"]
        store = s["store"]
        qty = s["qty"]

        row = asin_row.get(asin)
        if row is None:
            missing.append(f"  ASIN={asin} (发货表 {s['src_sheet']} Row{s['src_row']})")
            continue

        # 在采 -= 发货数量
        old_zc = _safe_int(ws.cell(row=row, column=zaicai_col).value)
        new_zc = old_zc - qty
        ws.cell(row=row, column=zaicai_col).value = new_zc

        # 根据发货店铺选择在途列
        if "稳再" in store:
            zt_col = wz_zt_col
            zt_label = "稳再在途"
        elif "柘流" in store:
            zt_col = zl_zt_col
            zt_label = "柘流在途"
        else:
            print(f"  [跳过] ASIN={asin} 发货店铺 '{store}' 无法识别为稳再/柘流")
            continue

        if zt_col is not None:
            old_zt = _safe_int(ws.cell(row=row, column=zt_col).value)
            ws.cell(row=row, column=zt_col).value = old_zt + qty

        updated += 1

        # 在采变负 → 报警
        if new_zc < 0:
            negative.append(
                f"  ASIN={asin} 在采={old_zc} → {new_zc} (扣减{qty}, 发货表 {s['src_sheet']} Row{s['src_row']})"
            )

    if updated > 0:
        wb.save(stats_path)
    wb.close()

    if missing:
        print(f"\n[警告] 以下 ASIN 在统计表中未找到 ({len(missing)} 条):")
        for m in missing:
            print(m)
    if negative:
        print(f"\n[报警] 以下 ASIN 在采数扣至负数 ({len(negative)} 条):")
        for n in negative:
            print(n)

    return {"updated": updated, "missing": len(missing), "negative": len(negative)}


# ── 工具函数 ──

def _safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(int(val)) if val == int(val) else str(val)
    return str(val).strip()


def _safe_int(val) -> int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0


# ── 主入口 ──

def cross_update(stats_path: str | Path, *shipping_paths: str | Path) -> dict:
    """跨表填写主函数。

    Returns:
        {updated, missing, negative, skipped, backup}
    """
    stats_path = Path(stats_path)
    tracker_path = stats_path.with_name(f"{stats_path.stem}.cross_track.json")
    tracker = _load_tracker(tracker_path)

    # 1. 读所有发货信息表（按日期去重）
    shipments: list[dict] = []
    total_skipped = 0
    for sp in shipping_paths:
        sp = Path(sp)
        print(f"读取发货信息表: {sp}")
        cutoff = tracker.get(sp.name)
        if cutoff:
            print(f"  上次处理日期: {cutoff}, 仅处理此日期之后的记录")
        batch, max_date = _read_shipping(sp, cutoff)
        print(f"  找到 {len(batch)} 条需处理的记录")
        if max_date:
            tracker[sp.name] = max_date
        shipments.extend(batch)

    if not shipments:
        print("未找到需处理的新数据 (所有记录均已处理或缺少有效数据)")
        return {"updated": 0, "missing": 0, "negative": 0, "skipped": 0}

    print(f"共 {len(shipments)} 条发货记录待处理")

    # 2. 备份统计表
    backup_path = stats_path.with_name(f"{stats_path.stem}_备份{stats_path.suffix}")
    shutil.copy2(stats_path, backup_path)
    print(f"已备份统计表: {backup_path}")

    # 3. 应用更新
    print(f"更新统计表: {stats_path}")
    result = _apply_updates(stats_path, shipments)
    result["backup"] = str(backup_path)

    # 4. 保存日期追踪（只在更新成功时写入）
    if result["updated"] > 0:
        _save_tracker(tracker_path, tracker)
        print(f"  已更新日期追踪: {tracker_path.name}")

    print(f"\n完成: 更新 {result['updated']} 行"
          f", 未找到 {result['missing']} 条"
          f", 负数报警 {result['negative']} 条")
    return result


def main():
    import argparse
    p = argparse.ArgumentParser(description="跨表数据填写：根据发货信息表更新统计表的在采/在途数量")
    p.add_argument("stats", help="统计表路径 (被更新的目标文件)")
    p.add_argument("shipping", nargs="+", help="发货信息表路径 (可多个, 数据来源只读)")
    p.add_argument("--json", action="store_true", help="输出结构化 JSON 运行摘要")
    args = p.parse_args()
    timer = RunTimer()
    try:
        result = cross_update(args.stats, *args.shipping)
        if args.json:
            print_json_summary({
                "module": "cross_table", "status": "ok",
                "elapsed": round(timer.elapsed, 1),
                "stats": str(args.stats),
                "shipping": [str(s) for s in args.shipping],
                **result,
            })
    except Exception as e:
        fatal_error("cross_table", e, stats=str(args.stats),
                     shipping=",".join(args.shipping))


if __name__ == "__main__":
    main()
