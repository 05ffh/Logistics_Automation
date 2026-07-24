"""读取发货明细表 Excel，按公司筛选行，按前缀提取物流单号。

列位通过第 2 行表头文字自动匹配，不再硬编码索引。
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

HEADER_ROW = 2  # 表头所在行

# 承运商前缀 → 公司名（长前缀优先，用于给单号归属公司并排序）
CARRIER_PREFIXES = [
    ("HYC", "华运昌"),
    ("HY", "华洋"),
    ("999", "云驼"),
    ("NZ", "宁致"),
    ("XM", "小满"),
]


def find_header_column(ws, header_name: str, row: int = HEADER_ROW) -> int | None:
    """按表头文字查找列索引（0-based），找不到返回 None。"""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=c).value
        if val and str(val).strip() == header_name:
            return c - 1
    return None


def identify_company(tn: str) -> str:
    """按前缀识别单号所属公司；未知承运商用前导字母/数字段作为分组键。"""
    for prefix, name in CARRIER_PREFIXES:
        if tn.startswith(prefix):
            return name
    m = re.match(r"^[A-Za-z]+", tn)
    return m.group(0) if m else tn[:3]


def company_position(all_tracking_nos: list[str], company_name: str) -> int:
    """公司在物流单号列中首次出现的次序（1-based），决定写入第几个物流轨迹列。"""
    order: list[str] = []
    for tn in all_tracking_nos:
        c = identify_company(tn)
        if c not in order:
            order.append(c)
    try:
        return order.index(company_name) + 1
    except ValueError:
        return 1


def find_company_rows(
    excel_path: str | Path,
    companies: list[dict],
) -> list[dict]:
    """扫描 Excel，返回所有公司的匹配行和对应单号。

    Args:
        companies: [{name, prefix}, ...]

    Returns:
        [{sheet, row_num, company, tracking_nos: [str], existing_info}]
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.strip().isdigit():
            continue
        ws = wb[sheet_name]
        # 按表头自动匹配列位，找不到回退到新规范默认值
        col_tracking_nos = find_header_column(ws, "物流单号")
        if col_tracking_nos is None:
            col_tracking_nos = 27  # AB列
        col_track_info = find_header_column(ws, "物流轨迹1")
        if col_track_info is None:
            col_track_info = 33  # AH列
        merged = _merged_value_map(ws)  # 合并单元格锚点值下传（如发货公司列常合并）
        for row_idx in range(3, ws.max_row + 1):
            tracking_str = _cell_str(ws, row_idx, col_tracking_nos, merged)
            if not tracking_str:
                continue
            existing = _cell_str(ws, row_idx, col_track_info, merged)

            # 按单号前缀归属公司（发货公司列填写不规范，前缀才是权威标识）
            for comp in companies:
                name = comp["name"]
                prefix = comp["prefix"]
                tns = _extract_by_prefix(tracking_str, prefix)
                if not tns:
                    continue
                results.append({
                    "sheet": sheet_name,
                    "row_num": row_idx,
                    "company": name,
                    "prefix": prefix,
                    "tracking_nos": tns,
                    "all_tracking_nos": _extract_all(tracking_str),
                    "existing_info": existing or None,
                })

    wb.close()
    return results


def _extract_by_prefix(text: str, prefix: str) -> list[str]:
    parts = re.split(r"[\n\r]+", text)
    seen = set()
    result = []
    for p in parts:
        p = p.strip()
        if p.startswith(prefix) and p not in seen:
            seen.add(p)
            result.append(p)
    return result


_TN_LINE = re.compile(r"^[A-Za-z0-9]{5,30}$")


def _extract_all(text: str) -> list[str]:
    """按物流单号列原始顺序提取所有单号（全公司），保序去重。"""
    result = []
    seen = set()
    for p in re.split(r"[\n\r]+", text):
        p = p.strip()
        if p and p not in seen and _TN_LINE.match(p):
            seen.add(p)
            result.append(p)
    return result


def _merged_value_map(ws) -> dict:
    """构建合并单元格查找表：区域内每个 (row, col) → 锚点(左上角)值。

    发货公司(K)等列常合并多行，openpyxl 只在锚点单元格保留值，
    其余读作 None。下传锚点值以便合并区内每一行都能正确匹配公司。
    """
    m: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
        if anchor is None:
            continue
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                m[(r, c)] = anchor
    return m


def _cell_str(ws, row: int, col: int, merged: dict | None = None) -> str:
    val = ws.cell(row=row, column=col + 1).value
    if val is None and merged is not None:
        val = merged.get((row, col + 1))
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(int(val)) if val == int(val) else str(val)
    return str(val).strip()
