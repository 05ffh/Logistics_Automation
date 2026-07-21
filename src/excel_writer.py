"""将物流轨迹信息写回 Excel 的"物流轨迹N"列。

每家发货公司独占一列：物流轨迹1(Y) / 物流轨迹2(Z) / 物流轨迹3 ...，
列号 N = 该公司单号在 S 列首次出现的次序。缺列时紧跟最后一个物流轨迹列后插入。
只写查询到的公司(云驼/宁致)对应列，不触碰其他公司(华洋/华运昌等)手填的列。

迁移清理：存量数据常把多家公司挤在物流轨迹1(Y)。写入查询公司到其正确列后，
把该公司单号的残留块从其他物流轨迹列移除（残留块的单号=已在正确列的新数据，可安全删）。
所有物流轨迹列统一采用物流轨迹1的列宽。
"""

from __future__ import annotations

import re
import shutil
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

HEADER_ROW = 2          # 表头所在行
TRACK_HEADER = "物流轨迹"  # 列标题前缀
_TN_LINE = re.compile(r"^[A-Za-z0-9]{5,30}$")


def write_results(excel_path: str | Path, results: list[dict]) -> dict:
    """按物流轨迹N列写回，写入前自动备份。

    每个 result 需含: sheet, row_num, routing_info, track_position；
    可选 tracking_nos（用于迁移清理其他列的残留块）。
    """
    excel_path = Path(excel_path)

    backup_path = excel_path.with_name(f"{excel_path.stem}_备份{excel_path.suffix}")
    shutil.copy2(excel_path, backup_path)

    try:
        with tempfile.NamedTemporaryFile(dir=excel_path.parent, delete=False) as f:
            f.write(b"lock_test")
        Path(f.name).unlink()
    except PermissionError:
        return {"updated": 0, "errors": 0, "locked": True}

    wb = openpyxl.load_workbook(excel_path)
    updated = 0
    errors = 0

    by_sheet: dict[str, list[dict]] = {}
    for r in results:
        by_sheet.setdefault(r["sheet"], []).append(r)

    for sheet_name, rows in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            errors += len(rows)
            continue

        ws = wb[sheet_name]
        need_max = max((r.get("track_position", 1) for r in rows), default=1)
        track_cols = _ensure_track_columns(ws, need_max)

        for r in rows:
            try:
                pos = r.get("track_position", 1)
                target = track_cols.get(pos)
                if target is None:
                    errors += 1
                    continue
                # 漏查不覆盖：与目标列旧内容按单号合并，本次没查到的单号保留旧块
                my_tns = r.get("tracking_nos") or []
                existing = ws.cell(row=r["row_num"], column=target).value
                merged = merge_preserve(r["routing_info"], str(existing or ""), my_tns)
                if not merged:
                    # 全部未查到且无旧数据 → 跳过，绝不用空值覆盖
                    continue
                ws.cell(row=r["row_num"], column=target).value = merged
                updated += 1
                _cleanup_other_columns(ws, r["row_num"], merged, track_cols, target)
            except Exception:
                errors += 1

        _normalize_sheet(ws)

    try:
        wb.save(excel_path)
    except (PermissionError, OSError) as e:
        wb.close()
        return {"updated": 0, "errors": errors, "backup": str(backup_path),
                "locked": True, "save_error": str(e)}
    wb.close()
    return {"updated": updated, "errors": errors, "backup": str(backup_path)}


def _cleanup_other_columns(ws, row_num, routing_info, track_cols, target_col):
    """迁移清理：把本次写入的单号残留块从其他物流轨迹列移除。"""
    written = _tns_in(routing_info)
    if not written:
        return
    for col in track_cols.values():
        if col == target_col:
            continue
        cell = ws.cell(row=row_num, column=col)
        if not cell.value:
            continue
        cleaned = _remove_blocks(str(cell.value), written)
        if cleaned != str(cell.value):
            cell.value = cleaned or None


def _tns_in(text: str) -> set[str]:
    return {ln.strip() for ln in text.split("\n") if _TN_LINE.match(ln.strip())}


def _parse_blocks(text: str) -> dict[str, str]:
    """把文本按单号行切成 {单号: "单号\\n轨迹..."} 块。"""
    blocks: dict[str, str] = {}
    cur_tn: str | None = None
    cur: list[str] = []
    for ln in text.split("\n"):
        if _TN_LINE.match(ln.strip()):
            if cur_tn is not None:
                blocks[cur_tn] = "\n".join(cur)
            cur_tn = ln.strip()
            cur = [ln]
        elif cur_tn is not None:
            cur.append(ln)
    if cur_tn is not None:
        blocks[cur_tn] = "\n".join(cur)
    return blocks


def merge_preserve(new_info: str, old_text: str, my_tns: list[str]) -> str:
    """按 my_tns 顺序重建：本次查到的用新块，没查到的保留旧块。"""
    new_b = _parse_blocks(new_info)
    old_b = _parse_blocks(old_text)
    out = []
    for tn in my_tns:
        if tn in new_b:
            out.append(new_b[tn])
        elif tn in old_b:
            out.append(old_b[tn])
    return "\n".join(out).strip()


def _remove_blocks(text: str, tns_to_remove: set[str]) -> str:
    """把文本按单号行切成块，删除单号命中的块，保留其余（含无单号的前导文本）。"""
    segments: list[tuple[str | None, list[str]]] = []
    cur_tn: str | None = None
    cur: list[str] = []
    for ln in text.split("\n"):
        if _TN_LINE.match(ln.strip()):
            segments.append((cur_tn, cur))
            cur_tn = ln.strip()
            cur = [ln]
        else:
            cur.append(ln)
    segments.append((cur_tn, cur))

    out: list[str] = []
    for tn, seg in segments:
        if tn is not None and tn in tns_to_remove:
            continue
        out.extend(seg)
    return "\n".join(out).strip()


def _ensure_track_columns(ws, need_max: int) -> dict[int, int]:
    """确保存在物流轨迹1..need_max列，返回 {N: 列号(1-based)}。

    缺失的列紧跟当前最后一个物流轨迹列之后插入并写表头；
    所有物流轨迹列统一采用物流轨迹1的列宽。
    """
    track_cols = find_track_columns(ws)
    if not track_cols:
        return track_cols

    existing_max = max(track_cols)
    base_col = track_cols.get(1, track_cols[existing_max])
    base_width = ws.column_dimensions[get_column_letter(base_col)].width

    if need_max > existing_max:
        insert_at = track_cols[existing_max] + 1   # 最后一个轨迹列的右侧
        amount = need_max - existing_max
        last_col = ws.max_column

        # insert_cols 会搬移单元格数据，但不搬列宽 → 先记录插入点右侧的列宽
        old_widths = {
            c: ws.column_dimensions[get_column_letter(c)].width
            for c in range(insert_at, last_col + 1)
        }
        ws.insert_cols(insert_at, amount)
        # 把右侧列宽整体右移 amount 列，对齐已搬移的数据
        for c in range(last_col, insert_at - 1, -1):
            w = old_widths.get(c)
            if w is not None:
                ws.column_dimensions[get_column_letter(c + amount)].width = w

        for n in range(existing_max + 1, need_max + 1):
            col = insert_at + (n - existing_max - 1)
            ws.cell(row=HEADER_ROW, column=col).value = f"{TRACK_HEADER}{n}"
            track_cols[n] = col

    # 所有物流轨迹列统一列宽 = 物流轨迹1
    if base_width:
        for col in track_cols.values():
            ws.column_dimensions[get_column_letter(col)].width = base_width

    return track_cols


def find_track_columns(ws) -> dict[int, int]:
    """扫描表头，返回已存在的 {N: 列号} for 物流轨迹N。"""
    cols: dict[int, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v and str(v).startswith(TRACK_HEADER):
            suffix = str(v)[len(TRACK_HEADER):].strip()
            if suffix.isdigit():
                cols[int(suffix)] = c
    return cols


def _normalize_sheet(ws) -> None:
    """统一数据行格式：行高自适应 + 内容列自动换行 + 轨迹列等线居中。

    修复业务 Excel 中 81 行起行高固定 22pt 导致多行内容被截断的问题，
    将所有数据行格式统一为 46~80 行的标准：行高自动、关键列 wrap、
    轨迹列字体等线 + 水平垂直居中。
    """
    from openpyxl.styles import Alignment, Font

    wrap_align = Alignment(wrap_text=True, vertical="top")
    track_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    track_font = Font(name="等线")

    content_cols, track_cols = _find_content_columns(ws)

    for row_idx in range(3, ws.max_row + 1):
        # 行高恢复自动
        rd = ws.row_dimensions.get(row_idx)
        if rd:
            rd.height = None

        # 内容列设置自动换行
        for col in content_cols:
            cell = ws.cell(row=row_idx, column=col)
            if cell.value and not cell.alignment.wrapText:
                cell.alignment = wrap_align

        # 轨迹列：等线字体 + 居中 + 自动换行
        for col in track_cols:
            cell = ws.cell(row=row_idx, column=col)
            if cell.value:
                cell.alignment = track_align
                cell.font = track_font


def _find_content_columns(ws) -> tuple[list[int], list[int]]:
    """返回 (内容列, 轨迹列)，根据表头识别。"""
    content_headers = {
        "成品编码", "sku", "备注", "货件号", "物流单号",
    }
    content_cols = []
    track_cols = []
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip()
        if h in content_headers:
            content_cols.append(c)
        if h.startswith("物流轨迹"):
            track_cols.append(c)
    return content_cols, track_cols
