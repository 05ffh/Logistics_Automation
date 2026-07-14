"""物流轨迹查询 - 主流程（多公司支持）。

用法:
    python -m src.main <excel_path>              # 处理所有 sheet + 所有公司
    python -m src.main <excel_path> 202605       # 只处理指定 sheet
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

# Windows 终端默认 GBK → 强制 stdout 使用 UTF-8，避免 print 中文时
# UnicodeEncodeError 崩溃。errors="replace" 兜底: 万一终端不认 UTF-8，
# 不可编码字符用 ? 替代而非抛异常（老 cmd.exe 可能仍乱码但不崩）。
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass

try:
    from .cdp_client import CdpClient
    from .companies.ningzhi import NingZhiAdapter
    from .companies.yuntuo import YunTuoAdapter
    from .companies.xiaoman import XiaoManAdapter
    from .excel_reader import find_company_rows, company_position
    from .excel_writer import write_results, merge_preserve, find_track_columns
    from .validation import is_valid_routing
    from .miss_tracker import (
        get_misses_path, load_misses, record_misses,
        remove_resolved, get_stubborn, print_miss_summary, MISS_THRESHOLD,
    )
except ImportError:
    from cdp_client import CdpClient
    from companies.ningzhi import NingZhiAdapter
    from companies.yuntuo import YunTuoAdapter
    from companies.xiaoman import XiaoManAdapter
    from excel_reader import find_company_rows, company_position
    from excel_writer import write_results, merge_preserve, find_track_columns
    from validation import is_valid_routing
    from miss_tracker import (
        get_misses_path, load_misses, record_misses,
        remove_resolved, get_stubborn, print_miss_summary, MISS_THRESHOLD,
    )

# 注册所有公司适配器
ADAPTERS = [NingZhiAdapter(), YunTuoAdapter(), XiaoManAdapter()]

# 异常检测阈值：查询数 >= N 且成功率 < RATE 判为疑似结构损坏 → 跳过写入保护存量
ANOMALY_MIN_COUNT = 5
ANOMALY_MIN_RATE = 0.5


def main():
    retry_stubborn = False
    args = [a for a in sys.argv[1:] if not a.startswith("--retry-stubborn")]
    if "--retry-stubborn" in sys.argv:
        retry_stubborn = True

    if not args:
        print("Usage: python -m src.main <excel_path> [sheet_names]")
        print("       python -m src.main <excel_path> --retry-stubborn")
        print("       python -m src.main --healthcheck")
        sys.exit(1)

    # 连接 CDP
    cdp_host = os.environ.get("CDP_HOST", "localhost:9222")
    host, _, port_str = cdp_host.partition(":")
    port = int(port_str) if port_str else 9222
    cdp = CdpClient(host=host, port=port)

    # 健康自检模式：用已知单号验证各站点结构是否还通
    if args[0] == "--healthcheck":
        print(f"Connecting to CDP at {host}:{port}...")
        try:
            cdp.list_tabs()
        except Exception:
            print("ERROR: Cannot reach Edge CDP. Is Edge running with --remote-debugging-port=9222?")
            sys.exit(1)
        ok = run_healthcheck(cdp)
        cdp.close()
        sys.exit(0 if ok else 1)

    excel_path = Path(args[0])
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    target_sheets = None
    if len(args) > 1:
        target_sheets = set(s.strip() for s in args[1].split(","))

    # 顽固补跑模式
    if retry_stubborn:
        run_retry_stubborn(excel_path, cdp, ADAPTERS)
        cdp.close()
        return

    # 1. 连接 CDP
    print(f"Connecting to CDP at {host}:{port}...")
    try:
        cdp.list_tabs()
    except Exception:
        print("ERROR: Cannot reach Edge CDP. Is Edge running with --remote-debugging-port=9222?")
        print("Double-click 物流网站一键启动.bat to start Edge first.")
        sys.exit(1)

    # 2. 读取 Excel
    print(f"Reading Excel: {excel_path}")
    company_configs = [{"name": a.name, "prefix": a.prefix} for a in ADAPTERS]
    all_rows = find_company_rows(excel_path, company_configs)
    if target_sheets:
        all_rows = [r for r in all_rows if r["sheet"] in target_sheets]

    if not all_rows:
        print("No matching rows found for any company.")
        cdp.close()
        return

    # 按公司分组
    by_company: dict[str, list[dict]] = {}
    for r in all_rows:
        by_company.setdefault(r["company"], []).append(r)

    adapter_map = {a.name: a for a in ADAPTERS}

    # 3. 逐公司查询
    all_results: dict[str, dict[str, str | None]] = {}  # {company: {tn: routing}}
    company_stats: dict[str, dict] = {}  # {company: {total, ok, suspect, reason}}

    for company_name, rows in by_company.items():
        adapter = adapter_map[company_name]
        print(f"\n{'='*40}")
        print(f"Company: {company_name} ({adapter.prefix}*)")

        # 切换到公司标签页
        ws_url = adapter.ensure_tab(cdp)
        cdp.connect_tab(ws_url)

        # 检查前置条件（登录/页面可达）
        if not adapter.check_ready(cdp):
            print(f"⚠️  WARNING: {company_name} 未就绪（可能未登录/页面不可达），跳过写入以保护存量。")
            company_stats[company_name] = {"total": 0, "ok": 0, "suspect": True,
                                           "reason": "未就绪"}
            continue

        # 收集唯一单号
        tns: list[str] = []
        seen_tns = set()
        for r in rows:
            for tn in r["tracking_nos"]:
                if tn not in seen_tns:
                    seen_tns.add(tn)
                    tns.append(tn)

        print(f"Found {len(rows)} rows, {len(tns)} unique {adapter.prefix}* numbers.")
        results = adapter.query(cdp, tns)
        res_map = {r.tracking_no: r.routing_info for r in results}
        all_results[company_name] = res_map

        # 异常检测：成功率过低 → 疑似页面结构变化/登录失效
        total = len(tns)
        ok = sum(1 for v in res_map.values() if v)
        suspect = total >= ANOMALY_MIN_COUNT and (ok / total) < ANOMALY_MIN_RATE
        company_stats[company_name] = {"total": total, "ok": ok,
                                       "suspect": suspect, "reason": "成功率异常"}
        if suspect:
            print(f"\n⚠️  WARNING: {company_name} 成功率异常 ({ok}/{total})，"
                  f"疑似页面结构变化或登录失效。")
            print(f"    为保护存量数据，本次跳过写入 {company_name}，请人工检查后重试。")

    # 疑似损坏的公司：不写入，保护存量数据
    suspect_companies = {c for c, s in company_stats.items() if s.get("suspect")}

    # 4. 合并结果并写入
    print(f"\n{'='*40}")
    print("Merging results...")
    updated_count = 0
    preserved_count = 0
    write_payload = []
    missed_entries: list[dict] = []

    # 跨所有公司的本次新查结果 {单号: 轨迹}
    global_results: dict[str, str] = {}
    for res in all_results.values():
        for tn, routing in res.items():
            if routing:
                global_results[tn] = routing

    for row in all_rows:
        company = row["company"]
        # 疑似损坏的公司跳过，绝不用可疑结果覆盖存量
        if company in suspect_companies:
            continue
        existing = row.get("existing_info") or ""
        old_map = _parse_existing_map(existing)
        # 仅本公司的单号（S 列顺序），每家公司独占一列
        my_tns = row.get("tracking_nos") or []

        blocks = []
        for tn in my_tns:
            fresh = global_results.get(tn)
            if fresh:
                blocks.append(f"{tn}\n{fresh}")
                updated_count += 1
            elif old_map.get(tn):
                blocks.append(f"{tn}\n{old_map[tn]}")
                preserved_count += 1
            else:
                # 本次没查到 + 旧也没有 → 真 MISS，记账
                missed_entries.append({
                    "company": company, "sheet": row["sheet"],
                    "row_num": row["row_num"], "tn": tn,
                    "my_tns": list(my_tns),
                })

        row["routing_info"] = "\n".join(blocks)
        # 该公司在 S 列首次出现的次序 → 写入第几个物流轨迹列
        pos = company_position(
            row.get("all_tracking_nos") or my_tns, company
        )
        row["track_position"] = pos
        # 补上该行该公司的 track_position
        for me in missed_entries:
            if (me["row_num"] == row["row_num"] and me["company"] == company
                    and "track_position" not in me):
                me["track_position"] = pos
        write_payload.append(row)

    # 5. 写回 Excel
    print(f"\nRouting: {updated_count} updated, {preserved_count} preserved.")
    print("Writing results back to Excel...")
    write_summary = write_results(excel_path, write_payload)

    if write_summary.get("locked"):
        print("ERROR: Excel file is open. Please close it and retry.")
    elif write_summary.get("backup"):
        print(f"Backup saved: {write_summary['backup']}")
    print(f"Done: {write_summary['updated']} rows updated, {write_summary['errors']} errors.")

    # 5.5 缺失单号追踪
    resolved = set(global_results.keys())
    removed = remove_resolved(excel_path, resolved)
    added = record_misses(excel_path, missed_entries)
    print_miss_summary(excel_path, new_count=added, removed_count=removed)

    # 6. 运行汇总（每公司成功率一目了然，便于发现悄然退化）
    print(f"\n{'='*40}")
    print("运行汇总:")
    for c, s in company_stats.items():
        rate = (s["ok"] / s["total"] * 100) if s["total"] else 0
        flag = "  ⚠️ 疑似异常，已跳过写入" if s.get("suspect") else ""
        print(f"  {c}: {s['ok']}/{s['total']} 成功 ({rate:.0f}%){flag}")
    if suspect_companies:
        print(f"\n⚠️ 有公司疑似异常({', '.join(suspect_companies)})，"
              f"已跳过写入保护存量数据，请人工核查页面/登录后重试。")

    cdp.close()


def run_healthcheck(cdp) -> bool:
    """金丝雀自检：用已知单号验证各站点结构是否还能正常抓取。

    任一公司拿不到合法轨迹，说明该站点可能改版/需重新登录 → 返回 False。
    """
    print(f"\n{'='*40}")
    print("健康自检 (canary)...")
    all_ok = True
    for adapter in ADAPTERS:
        canary = getattr(adapter, "canary_number", None)
        if not canary:
            print(f"  {adapter.name}: 跳过（未配置 canary 单号）")
            continue
        try:
            ws_url = adapter.ensure_tab(cdp)
            cdp.connect_tab(ws_url)
            if not adapter.check_ready(cdp):
                print(f"  {adapter.name}: FAIL — 未就绪（可能未登录）")
                all_ok = False
                continue
            # 自检重试几次，避免切换上下文后首查竞态导致的假告警
            routing = None
            for _ in range(3):
                results = adapter.query(cdp, [canary])
                routing = results[0].routing_info if results else None
                if is_valid_routing(routing):
                    break
            if is_valid_routing(routing):
                print(f"  {adapter.name}: PASS — {canary} 抓到合法轨迹")
            else:
                print(f"  {adapter.name}: FAIL — {canary} 未抓到合法轨迹（疑似结构变化）")
                all_ok = False
        except Exception as e:
            print(f"  {adapter.name}: FAIL — 异常 {type(e).__name__}: {e}")
            all_ok = False
    print(f"\n自检结果: {'全部通过 ✅' if all_ok else '存在失败 ❌，请检查对应站点'}")
    return all_ok


def run_retry_stubborn(excel_path, cdp, adapters):
    """--retry-stubborn 模式：只查顽固单号(miss_count>=2)，成功写回 + 移除。"""
    print("Retry Stubborn Mode")
    stubborn = get_stubborn(excel_path, threshold=MISS_THRESHOLD)
    if not stubborn:
        print("No stubborn numbers (need 2+ misses).")
        return

    by_company: dict[str, list[dict]] = {}
    for s in stubborn:
        by_company.setdefault(s["company"], []).append(s)

    print(f"\nFound {len(stubborn)} stubborn numbers across {len(by_company)} companies:")
    for c, entries in by_company.items():
        print(f"  {c}: {len(entries)} numbers")

    adapter_map = {a.name: a for a in adapters}
    total_resolved = 0
    total_still = 0
    all_writes: list[dict] = []       # (sheet, row_num, pos, tn, routing)

    for company_name, entries in by_company.items():
        adapter = adapter_map.get(company_name)
        if adapter is None:
            print(f"\nNo adapter for '{company_name}', skipping {len(entries)}.")
            continue

        print(f"\n{'='*40}")
        print(f"Company: {company_name}")

        ws_url = adapter.ensure_tab(cdp)
        cdp.connect_tab(ws_url)
        if not adapter.check_ready(cdp):
            print(f"  {company_name} not ready, skipping.")
            continue

        tns = [e["tn"] for e in entries]
        results = adapter.query(cdp, tns)
        res_map = {r.tracking_no: r.routing_info for r in results}

        resolved_tns = set()
        still_miss: list[dict] = []
        for e in entries:
            tn = e["tn"]
            routing = res_map.get(tn)
            if routing:
                resolved_tns.add(tn)
                total_resolved += 1
                all_writes.append({
                    "sheet": e["sheet"], "row_num": e["row_num"],
                    "company": company_name,
                    "track_position": e.get("track_position", 1),
                    "tn": tn, "my_tns": e.get("my_tns", [tn]),
                    "routing": routing,
                })
                print(f"  [{company_name}] {tn} RESOLVED")
            else:
                still_miss.append(e)
                total_still += 1
                print(f"  [{company_name}] {tn} STILL MISS (x{e.get('miss_count', 0)+1})")

        if resolved_tns:
            remove_resolved(excel_path, resolved_tns)
        if still_miss:
            record_misses(excel_path, still_miss)

    # 批量写回成功的顽固单号
    if all_writes:
        _write_stubborn_results(excel_path, all_writes)

    print(f"\n{'='*40}")
    print("Stubborn Retry Results:")
    print(f"  Resolved: {total_resolved}")
    print(f"  Still stuck: {total_still}")
    if total_still:
        print(f"  Re-run --retry-stubborn to retry again.")


def _write_stubborn_results(excel_path, writes: list[dict]):
    """批量写回顽固补查结果，按同一单元格合并，单次 open/save。"""
    import shutil
    import openpyxl

    backup = excel_path.with_name(f"{excel_path.stem}_备份{excel_path.suffix}")
    shutil.copy2(excel_path, backup)

    wb = openpyxl.load_workbook(excel_path)

    # 按 (sheet, row_num, track_position) 分组
    from collections import defaultdict
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for w in writes:
        groups[(w["sheet"], w["row_num"], w["track_position"])].append(w)

    for (sheet_name, row_num, pos), items in groups.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        track_cols = find_track_columns(ws)
        col = track_cols.get(pos)
        if col is None:
            continue

        existing = str(ws.cell(row=row_num, column=col).value or "")
        my_tns = items[0].get("my_tns") or [w["tn"] for w in items]
        # 新内容：每个成功的 tn + routing 块
        new_blocks = "\n".join(f"{w['tn']}\n{w['routing']}" for w in items)
        merged = merge_preserve(new_blocks, existing, my_tns)
        if merged:
            ws.cell(row=row_num, column=col).value = merged

    wb.save(excel_path)
    wb.close()
    print(f"\nWrote {len(writes)} stubborn results back to Excel.")


_TN_PATTERN = re.compile(r"^[A-Z0-9]{5,25}$")


def _extract_all_tns_from_s_column(y_text: str) -> list[str]:
    """从 Y 列文本中逐行提取所有物流单号（保持顺序）。"""
    tns = []
    if not y_text:
        return tns
    for line in y_text.split("\n"):
        stripped = line.strip()
        if _TN_PATTERN.match(stripped):
            tns.append(stripped)
    return tns


def _extract_routing_for_tn(y_text: str, tn: str) -> str:
    """从 Y 列提取指定单号的路由文本（不含单号行本身）。"""
    if not y_text or not tn:
        return ""
    all_tns = _extract_all_tns_from_s_column(y_text)
    try:
        idx = all_tns.index(tn)
    except ValueError:
        return ""
    # 找该单号在原文中的起始位置
    tn_start = y_text.find(tn)
    if tn_start == -1:
        return ""
    # 找下一个单号的位置作为结束边界
    if idx + 1 < len(all_tns):
        next_tn = all_tns[idx + 1]
        tn_end = y_text.find(next_tn, tn_start + len(tn))
    else:
        tn_end = len(y_text)
    routing = y_text[tn_start + len(tn) : tn_end].strip()
    return routing


def _parse_existing_map(y_text: str) -> dict[str, str]:
    """把旧 Y 列解析为 {单号: 轨迹文本}，用于保留其他公司/未刷新的轨迹。"""
    result: dict[str, str] = {}
    for tn in _extract_all_tns_from_s_column(y_text):
        routing = _extract_routing_for_tn(y_text, tn)
        if routing:
            result[tn] = routing
    return result


def _routing_equal(old: str, new: str) -> bool:
    """比较两条路由信息是否实质相同。"""
    return old.strip() == new.strip()


if __name__ == "__main__":
    main()
