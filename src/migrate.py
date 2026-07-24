"""将旧格式 Excel 迁移到 2026 发货信息表最终版规范。

通过表头名称自动匹配列位，不再硬编码列索引。
旧表头自动映射到新表头（如 成品编码→品名、发货渠道→实际发货渠道）。

用法:
  python -m src.migrate <旧文件路径> [--output <输出路径>]
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

HEADER_ROW = 2

# ── 新格式表头（37 列，与 最新格式.xlsx 一致）──────────────────
_NEW_HEADERS = [
    "发表日期",     # 1
    "图片",         # 2
    "品名",         # 3
    "发货店铺",     # 4
    "asin",         # 5
    "sku",          # 6
    "fnsku",        # 7
    "发货数量",     # 8
    "指定发货渠道", # 9
    "备注",         # 10
    "仓库",         # 11
    "箱数",         # 12
    "箱内数量",     # 13
    "箱规(长)",     # 14
    "箱规(宽)",     # 15
    "箱规(高)",     # 16
    "重量",         # 17
    "实际发货渠道", # 18
    "发货公司",     # 19
    "仓库发货时间", # 20
    "发车、发船时间", # 21
    "时效",         # 22
    "价格",         # 23
    "附加费",       # 24
    "是否自搬货",   # 25
    "条码确认",     # 26
    "货件号",       # 27
    "物流单号",     # 28
    "状态",         # 29
    "实发",         # 30
    "实发差值",     # 31
    "已收",         # 32
    "实收差值",     # 33
    "物流轨迹1",    # 34
    "物流轨迹2",    # 35
    "后台送达时段", # 36
    "更新日期",     # 37
]

NCOL = len(_NEW_HEADERS)

# ── 旧表头 → 新表头名称映射 ────────────────────────────────
_HEADER_RENAME: dict[str, str] = {
    "成品编码": "品名",
    "箱数量": "箱内数量",
    "发货渠道": "实际发货渠道",
    "预计发货时间": "仓库发货时间",
    "发船时间": "发车、发船时间",
    "配送时段": "时效",
    "发车、发船后配送时段": "时效",
    "发车。发船后提取时间": "时效",
}

# ── 统一样式 ────────────────────────────────────────────
_DATA_FONT = Font(name="等线", size=11)
_DATA_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_ALIGN = {10}  # J=备注
_IMAGE_COL = 2

# 需要做日期转换的列（旧值是 Excel 序列号或 datetime）
_DATE_HEADERS = {"发表日期", "仓库发货时间", "发车、发船时间", "更新日期"}


def _new_header_index(name: str) -> int | None:
    """返回新表头名称对应的列号 (1-based)，不存在返回 None。"""
    try:
        return _NEW_HEADERS.index(name) + 1
    except ValueError:
        return None


def _build_old_header_map(ws_old) -> dict[int, str]:
    """扫描旧表 Row 2，返回 {old_col_index: new_header_name}。

    旧 col1 无表头时自动识别为 发表日期。
    """
    mapping: dict[int, str] = {}
    for c in range(1, ws_old.max_column + 1):
        raw = str(ws_old.cell(row=HEADER_ROW, column=c).value or "").strip()
        if not raw:
            if c == 1:
                mapping[c] = "发表日期"  # 旧文件 A 列通常无表头，存日期
            continue
        # 旧名 → 新名
        new_name = _HEADER_RENAME.get(raw, raw)
        if _new_header_index(new_name) is not None:
            mapping[c] = new_name
    return mapping


# ── 主入口 ──────────────────────────────────────────────


def migrate(old_path: str | Path, output_path: str | Path | None = None) -> Path:
    old_path = Path(old_path)
    if output_path is None:
        output_path = old_path.with_name(f"{old_path.stem}_规范版{old_path.suffix}")
    else:
        output_path = Path(output_path)

    wb_old = openpyxl.load_workbook(old_path)
    wb_new = openpyxl.Workbook()
    first = True

    for sn in wb_old.sheetnames:
        ws_new = wb_new.active if first else wb_new.create_sheet(title=sn)
        if first:
            ws_new.title = sn
            first = False
        _migrate_sheet(wb_old[sn], ws_new)

    wb_new.save(output_path)
    wb_new.close()
    wb_old.close()
    print(f"Done: {output_path}")
    return output_path


# ── Sheet 迁移 ──────────────────────────────────────────


def _migrate_sheet(ws_old, ws_new) -> None:
    old_map = _build_old_header_map(ws_old)  # {old_col: new_header_name}
    col_letter_map = _build_col_letter_map(old_map)

    _write_headers(ws_new)
    _write_section_labels(ws_new)

    for r in range(3, ws_old.max_row + 1):
        _migrate_row(ws_old, ws_new, r, old_map, col_letter_map)

    _apply_data_format(ws_new)
    _auto_column_widths(ws_new)
    _auto_row_heights(ws_new)


def _build_col_letter_map(old_map: dict[int, str]) -> dict[str, str]:
    """构建旧列字母→新列字母映射，用于公式引用修正。"""
    mapping: dict[str, str] = {}
    for old_col, new_header in old_map.items():
        new_col = _new_header_index(new_header)
        if new_col:
            mapping[get_column_letter(old_col)] = get_column_letter(new_col)
    return mapping


def _fix_formula(formula: str, col_map: dict[str, str]) -> str:
    """修正公式中的列引用。'=I3*J3' → '=K3*L3'。"""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    # 匹配列引用：可选$，大写字母，后跟数字
    def _replace(m):
        col_letter = m.group(1)
        row_num = m.group(2)
        new_letter = col_map.get(col_letter, col_letter)
        return f"{new_letter}{row_num}"
    return re.sub(r"(\$?[A-Z]{1,3})(\d+)", _replace, formula)


def _migrate_row(ws_old, ws_new, old_row: int, old_map: dict[int, str],
                 col_letter_map: dict[str, str]) -> None:
    r = old_row
    for old_col, new_header in old_map.items():
        old_val = ws_old.cell(row=r, column=old_col).value
        if old_val is None:
            continue
        new_col = _new_header_index(new_header)
        if new_col is None:
            continue

        if new_header in _DATE_HEADERS:
            ws_new.cell(row=r, column=new_col, value=_convert_date(old_val))
        elif new_header == "价格":
            price, sur = _split_price(str(old_val))
            ws_new.cell(row=r, column=new_col, value=price)
            if sur != 0:
                sur_col = _new_header_index("附加费")
                if sur_col:
                    ws_new.cell(row=r, column=sur_col, value=sur)
        elif isinstance(old_val, str) and old_val.startswith("="):
            ws_new.cell(row=r, column=new_col,
                        value=_fix_formula(old_val, col_letter_map))
        else:
            ws_new.cell(row=r, column=new_col, value=old_val)


# ── 输出格式化 ──────────────────────────────────────────


def _write_headers(ws) -> None:
    hdr_font = Font(name="等线", size=11, bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill(patternType="solid", fgColor="FF5A5A5A")
    hdr_align = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(_NEW_HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = _DATA_BORDER
        cell.alignment = hdr_align


def _write_section_labels(ws) -> None:
    """动态计算分区合并范围：运营填写 A1~发货公司列，仓库填写 仓库发货时间列~末列。"""
    sec_font = Font(name="等线", size=11, bold=True, color="FFFFFFFF")
    sec_fill = PatternFill(patternType="solid", fgColor="FF5A5A5A")
    sec_align = Alignment(horizontal="center", vertical="center")

    ops_end = _new_header_index("发货公司")  # col 18
    wh_start = _new_header_index("仓库发货时间")  # col 19
    last = NCOL

    if ops_end:
        ops_letter = get_column_letter(ops_end)
        ws.merge_cells(f"A1:{ops_letter}1")
        c = ws.cell(row=1, column=1, value="运营填写")
        c.font = sec_font; c.fill = sec_fill; c.alignment = sec_align

    if wh_start:
        wh_start_letter = get_column_letter(wh_start)
        wh_end_letter = get_column_letter(last)
        ws.merge_cells(f"{wh_start_letter}1:{wh_end_letter}1")
        c = ws.cell(row=1, column=wh_start, value="仓库填写")
        c.font = sec_font; c.fill = sec_fill; c.alignment = sec_align


def _apply_data_format(ws) -> None:
    for r in range(3, ws.max_row + 1):
        for c in range(1, NCOL + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _DATA_FONT
            cell.border = _DATA_BORDER
            cell.alignment = _ALIGN_LEFT if c in _LEFT_ALIGN else _ALIGN_CENTER


def _auto_column_widths(ws) -> None:
    col_widths = {c: len(h) * 2.2 for c, h in enumerate(_NEW_HEADERS, 1)}
    for r in range(3, ws.max_row + 1):
        for c in range(1, NCOL + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            text = str(v)
            if text.startswith("=_xlfn.DISPIMG"):
                continue
            for line in text.split("\n"):
                w = sum(2 if ord(ch) > 127 else 1 for ch in line)
                if w > col_widths.get(c, 8):
                    col_widths[c] = w
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = min(w + 2, 40)


def _auto_row_heights(ws) -> None:
    col_w_units = {}
    for c in range(1, NCOL + 1):
        cl = get_column_letter(c)
        col_w_units[c] = (ws.column_dimensions[cl].width or 8)
    for r in range(3, ws.max_row + 1):
        max_lines = 1
        for c in range(1, NCOL + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or c == _IMAGE_COL:
                continue
            text = str(v)
            total = 0
            for line in text.split("\n"):
                char_w = sum(2 if ord(ch) > 127 else 1 for ch in line)
                cw = col_w_units.get(c, 8) or 8
                total += max(1, -(-char_w // int(cw)))
            if total > max_lines:
                max_lines = total
        ws.row_dimensions[r].height = max(max_lines * 15, 20)


# ── 数据转换 ────────────────────────────────────────────


def _convert_date(val) -> str | None:
    """Excel 序列号 / datetime / 'MMDD' → 'M月D日' 格式。"""
    if isinstance(val, datetime):
        return f"{val.month}月{val.day}日"
    if isinstance(val, (int, float)):
        dt = datetime(1899, 12, 30) + timedelta(days=int(val))
        return f"{dt.month}月{dt.day}日"
    s = str(val).strip()
    if re.match(r"^\d{4}$", s):
        return f"{int(s[:2])}月{int(s[2:])}日"
    return s


def _split_price(raw: str) -> tuple[float, float]:
    """'9+2' → (9.0, 2.0); '6.5' → (6.5, 0.0)"""
    m = re.match(r"^([\d.]+)\+(\d+)$", raw.strip())
    if m:
        return float(m.group(1)), float(m.group(2))
    try:
        return float(raw.strip()), 0.0
    except ValueError:
        return 0.0, 0.0


# ── CLI ─────────────────────────────────────────────────


def main():
    import argparse
    parser = argparse.ArgumentParser(description="迁移旧格式 Excel → 最终版规范")
    parser.add_argument("source", help="旧格式 Excel 路径")
    parser.add_argument("--output", "-o", default=None, help="输出路径")
    args = parser.parse_args()
    migrate(args.source, args.output)


if __name__ == "__main__":
    main()
