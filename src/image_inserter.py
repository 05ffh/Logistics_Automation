"""ASIN 图片自动匹配插入。

从映射源 Excel 提取 ASIN→图片库，然后为目标 Excel 的每行按 ASIN
自动嵌入对应产品图片到 B 列，使用 WPS cellImages + DISPIMG 机制。

用法:
  python -m src.image_inserter build <映射源Excel>     # 构建图片库
  python -m src.image_inserter insert <目标Excel>      # 插入图片
"""

from __future__ import annotations

import copy
import io
import os
import re
import shutil
import uuid
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ── 常量 ──────────────────────────────────────────────────────────

_IMAGE_COL = 2          # B 列 — 图片默认写入列（表头无"图片"时回退）
_ASIN_COL = 5           # E 列 — ASIN 默认来源列（表头无"asin"时回退）
_HEADER_ROW = 2
_DEFAULT_LIBRARY = "images/products"

# 表头匹配关键词（大小写不敏感）
_ASIN_HEADERS = {"asin"}
_IMAGE_HEADERS = {"图片"}

# WPS cellImages XML namespaces
_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
}

_CELLIMAGE_NS = "http://www.wps.cn/officeDocument/2017/etCustomData"
_IMAGE_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
)
_CELLIMAGE_REL_TYPE = "http://www.wps.cn/officeDocument/2020/cellImage"
_CONTENT_TYPE_OVERRIDE = (
    "application/vnd.wps-officedocument.cellimage+xml"
)

# Default cell image dimensions (EMU — ~7.5cm × ~7cm, fits typical cell)
_DEFAULT_CX = 1149350   # ~3.2cm — 参考测试格式.xlsx 的典型 cx
_DEFAULT_CY = 1149350  # ~3.2cm — 保持正方形，参考测试格式.xlsx
_OFF_X = 514350        # 参考测试格式.xlsx 的典型 offset
_ROW_HEIGHT_EMU = 952500


# ── build_library ─────────────────────────────────────────────────

def build_library(source_excel: str | Path,
                  output_dir: str | Path = _DEFAULT_LIBRARY) -> dict[str, str]:
    """从映射源 Excel 提取图片，按 ASIN 命名保存。返回 {ASIN: 文件名}。"""
    source_excel = Path(source_excel)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(source_excel) as z:
        id_to_rid = _parse_cellimages(z)
        rid_to_file = _parse_cellimages_rels(z)
        asin_to_id = _read_mapping_sheet(source_excel)

    mapping: dict[str, str] = {}
    for asin, img_id in asin_to_id.items():
        rid = id_to_rid.get(img_id)
        if not rid:
            print(f"  [WARN] ASIN {asin}: DISPIMG ID {img_id} not in cellimages.xml")
            continue
        img_path = rid_to_file.get(rid)
        if not img_path:
            print(f"  [WARN] ASIN {asin}: rId {rid} not in rels")
            continue

        with zipfile.ZipFile(source_excel) as z:
            img_data = z.read("xl/" + img_path)

        ext = Path(img_path).suffix
        dest = output_dir / f"{asin}{ext}"
        dest.write_bytes(img_data)
        mapping[asin] = dest.name
        print(f"  {asin} -> {dest.name} ({len(img_data)} bytes)")

    print(f"\nLibrary built: {len(mapping)} images -> {output_dir}")
    return mapping


def _parse_cellimages(z: zipfile.ZipFile) -> dict[str, str]:
    """从 cellimages.xml 提取 {DISPIMG_ID: rId}。"""
    if "xl/cellimages.xml" not in z.namelist():
        return {}
    xml = z.read("xl/cellimages.xml").decode("utf-8")
    result = {}
    for m in re.finditer(
        r'<etc:cellImage>.*?name="ID_([^"]+)".*?r:embed="(rId\d+)"',
        xml, re.DOTALL,
    ):
        result[m.group(1)] = m.group(2)
    return result


def _parse_cellimages_rels(z: zipfile.ZipFile) -> dict[str, str]:
    """从 cellimages.xml.rels 提取 {rId: media/imageN.ext}。"""
    if "xl/_rels/cellimages.xml.rels" not in z.namelist():
        return {}
    xml = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
    result = {}
    for m in re.finditer(r'Id="(rId\d+)".*?Target="([^"]+)"', xml):
        result[m.group(1)] = m.group(2)
    return result


def _read_mapping_sheet(source_excel: Path) -> dict[str, str]:
    """读取映射 Excel 的 sheet 数据，返回 {ASIN: DISPIMG_ID}。

    兼容两种格式:
      - 旧: A=ASIN, B=图片
      - 新: A=品名, B=asin, C=图片
    """
    wb = openpyxl.load_workbook(source_excel)
    ws = wb[wb.sheetnames[0]]

    # 按表头自动识别列位
    asin_col = 1
    img_col = 2
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip().lower()
        if h == "asin":
            asin_col = c
        elif h == "图片":
            img_col = c

    result = {}
    for r in range(2, ws.max_row + 1):
        asin = str(ws.cell(row=r, column=asin_col).value or "").strip()
        img = str(ws.cell(row=r, column=img_col).value or "")
        m = re.search(r'DISPIMG\("ID_([^"]+)"', img)
        if asin and m:
            result[asin] = m.group(1)
    wb.close()
    return result


# ── insert_images ─────────────────────────────────────────────────

def _fix_wps_shared_strings(f: Path) -> None:
    """WPS 文件在 Content_Types/rels 中声明了 sharedStrings.xml 但未写入 ZIP，
    openpyxl 会因找不到该文件而崩溃。仅在文件确实缺失时移除引用。"""
    import re as _re
    with zipfile.ZipFile(f) as zf:
        if "xl/sharedStrings.xml" in zf.namelist():
            return

    tmp = f.with_suffix(".wpsfix.zip")
    with zipfile.ZipFile(f) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = _re.sub(
                    r'<Override PartName="/xl/sharedStrings\.xml"[^>]*/>', "", text
                )
                data = text.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                text = _re.sub(r'<Relationship[^>]*sharedStrings[^>]*/>', "", text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(f)


def insert_images(target_excel: str | Path,
                  library_dir: str | Path = _DEFAULT_LIBRARY,
                  asin_col: int | None = None,
                  image_col: int | None = None,
                  sheet_names: list[str] | None = None,
                  ) -> dict:
    """为目标 Excel 按 ASIN 插入图片到指定列。返回 {updated, missing, skipped}。"""
    target_excel = Path(target_excel)
    library_dir = Path(library_dir)

    if not library_dir.is_dir():
        raise FileNotFoundError(f"Image library not found: {library_dir}")

    # 自动备份
    backup_path = target_excel.with_name(
        f"{target_excel.stem}_备份{target_excel.suffix}"
    )
    shutil.copy2(target_excel, backup_path)

    # 修复 WPS 文件：声明了 sharedStrings.xml 但未写入，openpyxl 会崩溃
    _fix_wps_shared_strings(target_excel)

    # 构建图片库索引 {ASIN: filename}
    lib: dict[str, str] = {}
    for f in library_dir.iterdir():
        if f.suffix.lower() in (".png", ".jpeg", ".jpg", ".gif", ".bmp"):
            lib[f.stem] = f.name

    # 扫描目标 Excel → 收集需要插入的 (row, ASIN, image_path)
    wb = openpyxl.load_workbook(target_excel)
    rows_by_asin: dict[str, list[tuple[str, int]]] = {}  # asin -> [(sheet, row), ...]
    missing = 0
    skipped = 0
    sheet_cols: dict[str, tuple[int, int]] = {}  # {sheet: (asin_col, image_col)}

    sheets = sheet_names if sheet_names else [
        sn for sn in wb.sheetnames
        if sn.isdigit() or not any(s.isdigit() for s in wb.sheetnames)
    ]

    for sn in sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        sc_asin = asin_col if asin_col else _find_header_column(ws, _ASIN_HEADERS, _ASIN_COL)
        sc_image = image_col if image_col else _find_header_column(ws, _IMAGE_HEADERS, _IMAGE_COL)
        sheet_cols[sn] = (sc_asin, sc_image)

        for r in range(3, ws.max_row + 1):
            asin = str(ws.cell(row=r, column=sc_asin).value or "").strip()
            if not asin:
                skipped += 1
                continue
            if asin not in lib:
                missing += 1
                continue
            rows_by_asin.setdefault(asin, []).append((sn, r))

    if not rows_by_asin:
        wb.close()
        return {"updated": 0, "missing": missing, "skipped": skipped,
                "backup": str(backup_path)}

    # 为每个唯一 ASIN 生成图片 ID + 分配 image 文件名
    unique_asins = sorted(rows_by_asin)
    img_entries: list[dict] = []  # [{id, rId, asin, img_path}]
    for i, asin in enumerate(unique_asins):
        img_file = library_dir / lib[asin]
        img_entries.append({
            "id": uuid.uuid4().hex.upper(),
            "rId": f"rId{i + 1}",
            "asin": asin,
            "img_path": img_file,
        })

    # ── Step 1: openpyxl 写 DISPIMG 公式 + 保存 ──
    id_by_asin = {e["asin"]: e["id"] for e in img_entries}
    img_by_asin = {e["asin"]: e["img_path"] for e in img_entries}
    updated = 0

    # 按 sheet 分组写入
    rows_for_sheet: dict[str, list[tuple[int, str]]] = {}
    for asin, locs in rows_by_asin.items():
        for sn, r in locs:
            rows_for_sheet.setdefault(sn, []).append((r, asin))

    for sn in sheets:
        if sn not in rows_for_sheet:
            continue
        ws = wb[sn]
        sc_asin, sc_image = sheet_cols[sn]
        col_letter = get_column_letter(sc_image)
        # 从目标文件继承图片列宽（若未设置过则用 8.5）
        cur_w = ws.column_dimensions[col_letter].width
        if not cur_w:
            ws.column_dimensions[col_letter].width = 8.5
            cur_w = 8.5
        col_px = cur_w * 7

        for r, asin in rows_for_sheet[sn]:
            ws.cell(row=r, column=sc_image).value = (
                f'=_xlfn.DISPIMG("ID_{id_by_asin[asin]}",1)'
            )
            rh = _image_row_height(img_by_asin[asin], col_px)
            ws.row_dimensions[r].height = rh
            updated += 1

    wb.save(target_excel)
    wb.close()

    # ── Step 2: ZIP 级写入 (openpyxl 保存后 cellImages 已丢失，从备份读取已有数据合并) ──
    _zip_write_images(target_excel, img_entries, rows_by_asin, backup_path)

    return {"updated": updated, "missing": missing, "skipped": skipped,
            "backup": str(backup_path)}


def _parse_existing_rels(zin: zipfile.ZipFile) -> tuple[list[str], int]:
    """解析已有 cellimages.xml.rels，返回 (Relationship 元素列表, 最大 rId 编号)。"""
    rels: list[str] = []
    max_rid = 0
    if "xl/_rels/cellimages.xml.rels" in zin.namelist():
        text = zin.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
        for m in re.finditer(r'<Relationship\s[^>]+/>', text):
            rels.append(m.group(0))
            rid_m = re.search(r'Id="rId(\d+)"', m.group(0))
            if rid_m:
                max_rid = max(max_rid, int(rid_m.group(1)))
    return rels, max_rid


def _parse_existing_cellimages(zin: zipfile.ZipFile) -> list[str]:
    """解析已有 cellimages.xml，返回 <etc:cellImage>...</etc:cellImage> 块列表。"""
    blocks: list[str] = []
    if "xl/cellimages.xml" not in zin.namelist():
        return blocks
    text = zin.read("xl/cellimages.xml").decode("utf-8")
    for m in re.finditer(r"<etc:cellImage>.*?</etc:cellImage>", text, re.DOTALL):
        blocks.append(m.group(0))
    return blocks


def _zip_write_images(target_excel: Path, img_entries: list[dict],
                      rows_by_asin: dict[str, list[tuple[str, int]]],
                      backup_path: Path) -> None:
    """ZIP 级写入: 合并已有 cellimages + 新增，保留其他 sheet 的图片数据。

    已有 cellImages 从 backup 读取（openpyxl save 已丢弃目标文件中的 cellImages）。
    """
    target_excel = Path(target_excel)
    backup_path = Path(backup_path)
    tmp_path = target_excel.with_suffix(".tmp")

    with zipfile.ZipFile(target_excel) as zin, \
         zipfile.ZipFile(backup_path) as zback, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        existing = set(zin.namelist())

        # ── 从备份读取已有 cellimages 数据（目标文件已被 openpyxl 清空）──
        old_rels_lines, max_rid = _parse_existing_rels(zback)
        old_cellimage_blocks = _parse_existing_cellimages(zback)

        # 从备份复制已有 media 文件，并计算 hash 用于去重
        existing_hashes: dict[str, int] = {}
        existing_media_nums: set[int] = set()
        for name in zback.namelist():
            m = re.match(r"xl/media/image(\d+)\.\w+", name)
            if m:
                num = int(m.group(1))
                existing_media_nums.add(num)
                data = zback.read(name)
                existing_hashes[_sha256_hex(data)] = num
        next_num = max(existing_media_nums, default=0)

        # ── 写入新图片（跳过 hash 重复的）──
        seen_hashes: dict[str, int] = {}
        for entry in img_entries:
            img_data = entry["img_path"].read_bytes()
            h = _sha256_hex(img_data)
            if h in existing_hashes:
                entry["media_num"] = existing_hashes[h]
            elif h in seen_hashes:
                entry["media_num"] = seen_hashes[h]
            else:
                next_num += 1
                entry["media_num"] = next_num
                seen_hashes[h] = next_num
                existing_hashes[h] = next_num
                ext = entry["img_path"].suffix
                zout.writestr(f"xl/media/image{next_num}{ext}", img_data)

        # ── 分配新 rId ──
        for i, entry in enumerate(img_entries):
            entry["rId"] = f"rId{max_rid + i + 1}"

        # ── 生成合并后的 cellimages.xml ──
        new_xml_blocks = _build_cellimages_xml_blocks(img_entries, rows_by_asin)
        all_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<etc:cellImages'
            ' xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
            ' xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">\n'
            + "\n".join(old_cellimage_blocks + new_xml_blocks)
            + "\n</etc:cellImages>"
        )
        zout.writestr("xl/cellimages.xml", all_xml.encode("utf-8"))

        # ── 生成合并后的 cellimages.xml.rels ──
        new_rels_lines = _build_cellimages_rels_lines(img_entries)
        rels_ns = 'xmlns="http://schemas.openxmlformats.org/package/2006/relationships"'
        all_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<Relationships {rels_ns}>\n'
            + "\n".join(old_rels_lines + new_rels_lines)
            + "\n</Relationships>"
        )
        zout.writestr("xl/_rels/cellimages.xml.rels", all_rels.encode("utf-8"))

        # ── 复制所有文件：目标（openpyxl 输出）+ 备份（cellImages / media）──
        new_media_nums = {e["media_num"] for e in img_entries if e["media_num"] not in existing_media_nums}
        written: set[str] = set()

        # 先处理目标文件（不含 cellImages 相关内容）
        for name in sorted(existing):
            if name == "xl/cellimages.xml" or name == "xl/_rels/cellimages.xml.rels":
                continue
            if name.startswith("xl/media/"):
                continue  # 目标文件无 media，跳过
            data = zin.read(name)
            if name == "[Content_Types].xml":
                data = _update_content_types(data)
            elif name == "xl/_rels/workbook.xml.rels":
                data = _update_workbook_rels(data)
            zout.writestr(name, data)
            written.add(name)

        # 再从备份复制 cellImages 相关文件（media + Content_Types + workbook.rels 需要合并）
        for name in sorted(zback.namelist()):
            if name in written:
                continue
            if name == "xl/cellimages.xml" or name == "xl/_rels/cellimages.xml.rels":
                continue  # 已生成合并版本
            if name.startswith("xl/media/"):
                m = re.match(r"xl/media/image(\d+)\.\w+", name)
                if m and int(m.group(1)) in new_media_nums:
                    continue  # 新图片已在上面写入
                zout.writestr(name, zback.read(name))
            elif name == "[Content_Types].xml":
                data = _update_content_types(zback.read(name))
                zout.writestr(name, data)
            elif name == "xl/_rels/workbook.xml.rels":
                data = _update_workbook_rels(zback.read(name))
                zout.writestr(name, data)
            else:
                # 其他备份独有的文件（如 Sheet3 相关）
                zout.writestr(name, zback.read(name))

    tmp_path.replace(target_excel)


def _build_cellimages_xml_blocks(entries: list[dict],
                                  rows_by_asin: dict[str, list[tuple[str, int]]]) -> list[str]:
    """生成新的 <etc:cellImage> 块列表。"""
    blocks = []
    c_nv_pr_id = 2 + len(entries)  # 避免 ID 冲突（简单递增）
    for entry in entries:
        c_nv_pr_id += 1
        locs = rows_by_asin.get(entry["asin"], [])
        first_row = min(r for _, r in locs) if locs else 3
        y_off = (first_row - 2) * _ROW_HEIGHT_EMU

        block = (
            '<etc:cellImage>'
            '<xdr:pic>'
            '<xdr:nvPicPr>'
            f'<xdr:cNvPr id="{c_nv_pr_id}" name="ID_{entry["id"]}"/>'
            '<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
            '</xdr:nvPicPr>'
            '<xdr:blipFill>'
            f'<a:blip r:embed="{entry["rId"]}"/>'
            '<a:stretch><a:fillRect/></a:stretch>'
            '</xdr:blipFill>'
            '<xdr:spPr>'
            '<a:xfrm>'
            f'<a:off x="{_OFF_X}" y="{y_off}"/>'
            f'<a:ext cx="{_DEFAULT_CX}" cy="{_DEFAULT_CY}"/>'
            '</a:xfrm>'
            '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            '<a:noFill/>'
            '<a:ln w="9525"><a:noFill/></a:ln>'
            '</xdr:spPr>'
            '</xdr:pic>'
            '</etc:cellImage>'
        )
        blocks.append(block)
    return blocks


def _build_cellimages_rels_lines(entries: list[dict]) -> list[str]:
    """生成新的 Relationship 行列表（不含 XML 头和闭合标签）。"""
    lines = []
    for entry in entries:
        ext = entry["img_path"].suffix
        lines.append(
            f'<Relationship Id="{entry["rId"]}"'
            f' Type="{_IMAGE_REL_TYPE}"'
            f' Target="media/image{entry["media_num"]}{ext}"/>'
        )
    return lines


def _update_content_types(data: bytes) -> bytes:
    """确保 [Content_Types].xml 注册 image/jpeg, image/png 和 cellimages Override。"""
    text = data.decode("utf-8")

    if "image/png" not in text:
        text = text.replace(
            "</Types>",
            '<Default Extension="png" ContentType="image/png"/></Types>',
        )
    if "image/jpeg" not in text:
        text = text.replace(
            "</Types>",
            '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>',
        )
    if "cellimages.xml" not in text:
        text = text.replace(
            "</Types>",
            f'<Override PartName="/xl/cellimages.xml" ContentType="{_CONTENT_TYPE_OVERRIDE}"/></Types>',
        )

    return text.encode("utf-8")


def _update_workbook_rels(data: bytes) -> bytes:
    """确保 workbook.xml.rels 包含 cellImage 关系。"""
    text = data.decode("utf-8")
    if "cellImage" in text:
        return data

    # 找最大 rId 编号
    nums = [int(m.group(1)) for m in re.finditer(r'Id="rId(\d+)"', text)]
    next_rid = max(nums, default=0) + 1

    text = text.replace(
        "</Relationships>",
        f'<Relationship Id="rId{next_rid}" Type="{_CELLIMAGE_REL_TYPE}"'
        ' Target="cellimages.xml"/></Relationships>',
    )
    return text.encode("utf-8")


def _find_header_column(ws, headers: set[str], fallback: int) -> int:
    """扫描表头行，按关键词匹配列位（大小写不敏感）。未找到则返回 fallback。"""
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=_HEADER_ROW, column=c).value or "").strip().lower()
        if v in {h.lower() for h in headers}:
            return c
    return fallback


def _sha256_hex(data: bytes) -> str:
    import hashlib
    return hashlib.sha256(data).hexdigest()


def _image_row_height(img_path: Path, col_px: float) -> float:
    """根据图片宽高比和列像素宽度，计算自适应行高 (pt)。"""
    w, h = _image_dimensions(img_path)
    if w == 0 or h == 0:
        return 100  # 默认
    # 行高 = (图片高度 / 图片宽度) * 列宽 * 0.75 (px→pt 换算)
    return round((h / w) * col_px * 0.75, 1)


def _image_dimensions(path: Path) -> tuple[int, int]:
    """读取 PNG/JPEG 图片尺寸，不依赖第三方库。"""
    data = path.read_bytes()
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        # PNG: width at bytes 16-19, height at 20-23 (big-endian)
        w = int.from_bytes(data[16:20], "big")
        h = int.from_bytes(data[20:24], "big")
        return w, h
    if data[:2] == b"\xff\xd8":
        # JPEG: scan for SOF0/SOF2 marker
        i = 2
        while i < len(data) - 9:
            if data[i] != 0xFF:
                break
            marker = data[i + 1]
            if marker in (0xC0, 0xC2):  # SOF0, SOF2
                h = int.from_bytes(data[i + 5:i + 7], "big")
                w = int.from_bytes(data[i + 7:i + 9], "big")
                return w, h
            i += 2 + int.from_bytes(data[i + 2:i + 4], "big")
    return 0, 0


# ── CLI ───────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="ASIN 图片自动匹配")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_build = sub.add_parser("build", help="从映射源 Excel 构建图片库")
    p_build.add_argument("source", help="映射源 Excel 路径")
    p_build.add_argument("--output", "-o", default=_DEFAULT_LIBRARY,
                         help=f"图片库输出目录 (默认: {_DEFAULT_LIBRARY})")

    p_insert = sub.add_parser("insert", help="为目标 Excel 插入图片")
    p_insert.add_argument("target", help="目标 Excel 路径")
    p_insert.add_argument("--library", "-l", default=_DEFAULT_LIBRARY,
                          help=f"图片库目录 (默认: {_DEFAULT_LIBRARY})")
    p_insert.add_argument("--asin-col", default=None,
                          help="ASIN 所在列 (默认: 按表头自动识别，回退 E)")
    p_insert.add_argument("--image-col", default=None,
                          help="图片写入列 (默认: 按表头自动识别，回退 B)")
    p_insert.add_argument("--sheet", "-s", action="append", default=None,
                          help="只处理指定 sheet (可多次指定)")

    args = parser.parse_args()

    if args.cmd == "build":
        result = build_library(args.source, args.output)
        print(f"\nDone. {len(result)} images extracted.")

    elif args.cmd == "insert":
        asin_col = openpyxl.utils.column_index_from_string(args.asin_col) if args.asin_col else None
        image_col = openpyxl.utils.column_index_from_string(args.image_col) if args.image_col else None
        result = insert_images(args.target, args.library,
                               asin_col=asin_col, image_col=image_col,
                               sheet_names=args.sheet)
        print(f"\nDone. Updated: {result['updated']}, "
              f"Missing (ASIN not in library): {result['missing']}, "
              f"Skipped (empty ASIN): {result['skipped']}")
        print(f"Backup: {result['backup']}")


if __name__ == "__main__":
    main()
